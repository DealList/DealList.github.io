"""DCM sheet.xlsx 포맷의 신규 엑셀 생성.

전략: 원본 DCM sheet을 템플릿으로 로드 → 모든 데이터 행 삭제 → 헤더만 남김 →
신규 트랜치 데이터 채워넣기. 헤더 서식·열 너비·폰트·정렬은 자동 보존됨.
"""
from __future__ import annotations
import shutil
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection
from copy import copy
import json
import zipfile
import tempfile

import config
from parser import TrancheRecord
from formulas import build_lead_formula


def _copy_cell_style(src, dst):
    """openpyxl cell 서식 복사."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _find_group_merges(ws):
    """row 1 에서 '주관' 과 '인수' 그룹 병합 위치 찾기. 못 찾으면 (None, None)."""
    lead_mc = uw_mc = None
    for mc in ws.merged_cells.ranges:
        if mc.min_row != 1 or mc.max_row != 1:
            continue
        v = ws.cell(1, mc.min_col).value
        if v == "주관":
            lead_mc = (mc.min_col, mc.max_col)
        elif v == "인수":
            uw_mc = (mc.min_col, mc.max_col)
    return lead_mc, uw_mc


def _insert_cols_safe(ws, insert_at: int, count: int) -> None:
    """병합 셀을 안전하게 처리하면서 컬럼 삽입.

    openpyxl insert_cols 가 병합 셀 좌표를 자동 갱신하지 않는 경우가 있어
    모든 병합을 미리 해제하고, 좌표를 수동으로 시프트한 뒤 다시 병합한다.
    """
    saved = [(mc.min_row, mc.min_col, mc.max_row, mc.max_col)
             for mc in list(ws.merged_cells.ranges)]
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    ws.insert_cols(insert_at, amount=count)

    for r1, c1, r2, c2 in saved:
        if c1 >= insert_at:
            c1 += count
            c2 += count
        elif c2 >= insert_at:
            c2 += count
        ws.merge_cells(start_row=r1, end_row=r2,
                       start_column=c1, end_column=c2)


def _expand_broker_columns(ws) -> None:
    """mappings 의 broker 리스트가 시트 헤더보다 길면 헤더 컬럼 추가.

    - 주관 영역에 새 alias: 주관 끝(=인수 시작 직전)에 컬럼 삽입 → 인수 영역 시프트
    - 인수 영역에 새 alias: 인수 끝 다음에 alias 입력 + 인수 그룹 병합 확장
    """
    leads = config.LEAD_MANAGERS
    uws = config.UNDERWRITERS

    lead_mc, uw_mc = _find_group_merges(ws)
    if lead_mc is None or uw_mc is None:
        return

    sheet_lead_count = lead_mc[1] - lead_mc[0] + 1
    sheet_uw_count = uw_mc[1] - uw_mc[0] + 1
    new_lead_count = max(0, len(leads) - sheet_lead_count)
    new_uw_count = max(0, len(uws) - sheet_uw_count)

    if new_lead_count == 0 and new_uw_count == 0:
        return

    # 1) 주관 추가: insert_cols 로 인수 영역 모두 시프트
    if new_lead_count > 0:
        insert_at = lead_mc[1] + 1
        _insert_cols_safe(ws, insert_at, new_lead_count)
        # 주관 그룹 병합 확장
        ws.unmerge_cells(start_row=1, start_column=lead_mc[0],
                         end_row=1, end_column=lead_mc[1])
        new_lead_max = lead_mc[1] + new_lead_count
        ws.merge_cells(start_row=1, start_column=lead_mc[0],
                       end_row=1, end_column=new_lead_max)
        # 새 주관 alias + 헤더 서식
        for i in range(new_lead_count):
            alias = leads[sheet_lead_count + i]
            cell = ws.cell(row=2, column=insert_at + i, value=alias)
            _copy_cell_style(ws.cell(row=2, column=lead_mc[1]), cell)
        # 인수 영역 좌표 시프트 (이후 인수 처리에 사용)
        uw_mc = (uw_mc[0] + new_lead_count, uw_mc[1] + new_lead_count)

    # 2) 인수 추가: 끝에 alias write + 그룹 병합 확장
    if new_uw_count > 0:
        ws.unmerge_cells(start_row=1, start_column=uw_mc[0],
                         end_row=1, end_column=uw_mc[1])
        new_uw_max = uw_mc[1] + new_uw_count
        ws.merge_cells(start_row=1, start_column=uw_mc[0],
                       end_row=1, end_column=new_uw_max)
        for i in range(new_uw_count):
            alias = uws[sheet_uw_count + i]
            cell = ws.cell(row=2, column=uw_mc[1] + 1 + i, value=alias)
            _copy_cell_style(ws.cell(row=2, column=uw_mc[1]), cell)
        # row 1 새 셀 서식 (병합되어 보이지 않지만 셀 단위 서식은 보존)
        src1 = ws.cell(row=1, column=uw_mc[0])
        for col in range(uw_mc[1] + 1, new_uw_max + 1):
            _copy_cell_style(src1, ws.cell(row=1, column=col))

    # 신규 추가된 broker 컬럼들 width 표준값 적용 (다른 broker 컬럼과 통일)
    # XML 5.0 → Excel UI 4.42 (= 60 px)
    if new_lead_count > 0 or new_uw_count > 0:
        from openpyxl.utils import get_column_letter
        # 주관 신규 영역
        if new_lead_count > 0:
            for c in range(insert_at, insert_at + new_lead_count):
                ws.column_dimensions[get_column_letter(c)].width = 5.0
        # 인수 신규 영역 (uw_mc 는 위에서 시프트된 좌표)
        if new_uw_count > 0:
            for c in range(uw_mc[1] + 1, new_uw_max + 1):
                ws.column_dimensions[get_column_letter(c)].width = 5.0


def _normalize_borders(ws) -> None:
    """모든 셀 테두리를 thin + 라이트 그레이 (Excel 기본 격자선 색) 로 통일.

    사용자가 _ui_test.xlsx 에 직접 적용한 색 (theme=0, tint=-0.15 — 옅은 회색) 을
    그대로 채택. medium/thick/double 및 굵은 외곽 모두 일관된 thin 회색으로 정규화.

    범위 결정: ws.max_row/max_column 만 보면 데이터보다 한참 아래쪽에 남아있는
    원본 템플릿의 잔재 (예: row 290 의 겹줄) 를 놓침. ws._cells 의 실제 styled 셀
    좌표 전부 스캔해서 진짜 max 좌표 얻고 + 충분한 버퍼 (+500행, +20열).
    """
    light_gray = Color(theme=0, tint=-0.14996795556505021)
    thin_side = Side(style="thin", color=light_gray)
    uniform_border = Border(left=thin_side, right=thin_side,
                            top=thin_side, bottom=thin_side)

    # ws._cells 는 메모리에 존재하는 모든 cell 의 dict (key=(row,col)).
    # 헤더 셀 + 데이터 셀 + 템플릿에서 남은 styled-but-empty 셀 모두 포함.
    real_max_row = ws.max_row or 1
    real_max_col = ws.max_column or 1
    try:
        for (r, c) in ws._cells.keys():
            if r > real_max_row:
                real_max_row = r
            if c > real_max_col:
                real_max_col = c
    except AttributeError:
        pass

    # 사용 영역 + 충분한 여유분
    max_row = real_max_row + 500
    max_col = real_max_col + 20
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = uniform_border


def _add_ignored_errors_xml(xlsx_path: Path) -> None:
    """저장된 xlsx 의 각 worksheet XML 에 <ignoredErrors> 요소 추가.

    openpyxl 3.x 가 ignored_errors 속성을 지원 안 해서 (Worksheet 클래스에 attribute
    자체 없음) xlsx 압축 풀어 XML 직접 수정. 회차 컬럼(C) 의 '16-1' 같은 텍스트가
    숫자 오인 경고 (녹색 삼각형) 띄우는 것을 sheet 별로 numberStoredAsText="1" 로 무시.
    """
    # 회차 컬럼의 '16-1' / '45-2' 같이 '-' 포함 텍스트가 Excel 에서 여러 종류의
    # 오류 (numberStoredAsText, evalError, twoDigitTextYear, formulaRange 등)로
    # 표시될 수 있어 모든 가능한 종류를 한 번에 무시.
    ignored_xml = (
        '<ignoredErrors>'
        '<ignoredError sqref="C3:C10000" '
        'numberStoredAsText="1" '
        'evalError="1" '
        'twoDigitTextYear="1" '
        'formula="1" '
        'formulaRange="1" '
        'unlockedFormula="1" '
        'emptyCellReference="1" '
        'listDataValidation="1" '
        'calculatedColumn="1"/>'
        '</ignoredErrors>'
    )
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        with zipfile.ZipFile(xlsx_path, "r") as z:
            z.extractall(tmp_path)

        sheets_dir = tmp_path / "xl" / "worksheets"
        if sheets_dir.exists():
            for sheet_file in sorted(sheets_dir.glob("sheet*.xml")):
                content = sheet_file.read_text(encoding="utf-8")
                if "<ignoredErrors>" in content:
                    continue
                # </worksheet> 직전에 삽입
                content = content.replace("</worksheet>", ignored_xml + "</worksheet>")
                sheet_file.write_text(content, encoding="utf-8")

        out_tmp = xlsx_path.with_suffix(".rebuild.xlsx")
        with zipfile.ZipFile(out_tmp, "w", zipfile.ZIP_DEFLATED) as z:
            for f in tmp_path.rglob("*"):
                if f.is_file():
                    z.write(f, f.relative_to(tmp_path))
        shutil.move(str(out_tmp), str(xlsx_path))


def _build_template(output_path: Path) -> None:
    """템플릿 복사 후 데이터 행(3행 이하)을 모두 삭제 + broker 컬럼 동적 확장 + UI 정리."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(config.TEMPLATE_XLSX, output_path)

    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 병합된 데이터 영역 셀 해제 (행1~2 헤더 병합은 보존)
        for mc in list(ws.merged_cells.ranges):
            if mc.min_row >= 3:
                ws.unmerge_cells(str(mc))

        # 행3부터 끝까지 삭제 (값+서식 모두)
        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)

        # row_dimensions 잔재 제거: delete_rows 는 ws.max_row 이내만 정리하므로
        # 템플릿에 우연히 남은 row_dimensions (예: 옛 편집 시 hidden=True 잔재) 가
        # 그대로 살아남아 데이터 row 의 표시 속성을 덮어쓰는 사고가 있었음
        # (2024년 605행 뉴스테이허브제1호 2회차 사례). 행3 이상의 dimension 전부 purge.
        rd_keys_to_remove = [r for r in ws.row_dimensions if r >= 3]
        for r in rd_keys_to_remove:
            del ws.row_dimensions[r]

        # mappings 에 새로 추가된 broker 가 있으면 헤더 컬럼 확장
        _expand_broker_columns(ws)

        # 사용자 요청 UI 정리:
        # (a) 모든 셀 테두리를 thin 단일 스타일로 통일 (엑셀 기본 셀테두리 형태)
        # (b) 1~2행 헤더 상단 고정 (스크롤해도 헤더 보이도록).
        #     원본 템플릿이 마지막 편집 위치(예: A151) 에 sheet_view 가 저장되어 있어
        #     freeze 만 설정하면 그 위치에서 시작. topLeftCell + selection 도 A3 으로 reset.
        _normalize_borders(ws)
        ws.freeze_panes = "A3"
        ws.sheet_view.topLeftCell = "A1"
        if ws.sheet_view.pane is not None:
            ws.sheet_view.pane.topLeftCell = "A3"
        ws.sheet_view.selection = [Selection(activeCell="A3", sqref="A3")]

    wb.save(output_path)


def _data_row_alignment_center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _data_row_alignment_right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _font() -> Font:
    return Font(name="맑은 고딕", size=11)


def _write_tranche_row(ws, row_idx: int, rec: TrancheRecord) -> None:
    """1개 트랜치를 1행에 기록."""
    f = _font()
    align_c = _data_row_alignment_center()
    align_r = _data_row_alignment_right()

    def put(col, val, *, align=align_c, fmt=None):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = f
        cell.alignment = align
        if fmt:
            cell.number_format = fmt

    put(config.COL["청약일"], rec.subscription_date, fmt="yyyy/mm/dd;@")
    put(config.COL["발행사"], rec.issuer_alias)
    put(config.COL["회차"], rec.series)
    put(config.COL["종류"], rec.bond_type)
    put(config.COL["신용등급"], rec.credit_rating)

    if isinstance(rec.maturity, (date, datetime)):
        put(config.COL["만기일"], rec.maturity, fmt="yyyy/mm/dd;@")
    else:
        put(config.COL["만기일"], rec.maturity or "")

    # 액수 표시는 반올림 (소수점 없이). meta 의 raw value 는 유지 (보강 시점에 정확도).
    # 단 Excel 표시는 사용자 요청대로 무조건 round (엠지캐피탈 4-2 류 273.34 → 273).
    def _round_amt(v):
        return round(v) if isinstance(v, (int, float)) else v
    put(config.COL["최초모집"], _round_amt(rec.initial_amount), align=align_r)
    put(config.COL["발행한도"], _round_amt(rec.issue_limit), align=align_r)
    put(config.COL["수요예측"], _round_amt(rec.demand_amount), align=align_r)
    put(config.COL["최종발행"], _round_amt(rec.final_amount), align=align_r)
    put(config.COL["회차합산"], _round_amt(rec.series_total), align=align_r)

    # 경쟁률 = 수요예측 / 최초모집 (수식). demand=0 도 표시 (수요예측 참여 0건 케이스)
    if rec.demand_amount is not None and rec.initial_amount:
        col_letter_I = get_column_letter(config.COL["수요예측"])
        col_letter_G = get_column_letter(config.COL["최초모집"])
        formula = f"={col_letter_I}{row_idx}/{col_letter_G}{row_idx}"
        put(config.COL["경쟁률"], formula, align=align_r, fmt="0.00")

    put(config.COL["희망금리"], rec.rate_target)
    put(config.COL["수요금리"], rec.rate_demand)
    if rec.rate_final is not None:
        put(config.COL["최종금리"], rec.rate_final, fmt="0.000")

    # 인수 AO~BT (공시 그대로). 액수 반올림 (소수점 없이 표시).
    uw_start = config.underwriter_col_start()
    for i, alias in enumerate(config.UNDERWRITERS):
        amt = rec.underwriter_alloc.get(alias)
        if amt:
            put(uw_start + i, _round_amt(amt), align=align_r)

    # 주관 P~AN: 대표 증권사 셀에만 산식 박기
    for alias in rec.lead_managers:
        if alias not in config.LEAD_MANAGERS:
            continue
        col = config.COL["주관_시작"] + config.LEAD_MANAGERS.index(alias)
        formula = build_lead_formula(row_idx, alias, rec.lead_managers)
        if formula:
            put(col, formula, align=align_r, fmt="0")


def _record_to_dict(r: TrancheRecord) -> dict:
    """TrancheRecord 를 JSON 호환 dict 로 직렬화 (meta.json 보관용)."""
    def ser_date(v):
        if isinstance(v, (date, datetime)):
            return v.isoformat()[:10]
        return v
    return {
        "subscription_date": ser_date(r.subscription_date),
        "issuer_alias": r.issuer_alias,
        "issuer_full": r.issuer_full,
        "corp_code": r.corp_code,
        "series": r.series,
        "bond_type": r.bond_type,
        "credit_rating": r.credit_rating,
        "maturity": ser_date(r.maturity),
        "initial_amount": r.initial_amount,
        "issue_limit": r.issue_limit,
        "demand_amount": r.demand_amount,
        "final_amount": r.final_amount,
        "series_total": r.series_total,
        "rate_target": r.rate_target,
        "rate_demand": r.rate_demand,
        "rate_final": r.rate_final,
        "lead_managers": list(r.lead_managers),
        "underwriter_alloc": dict(r.underwriter_alloc),
        "rcept_no": r.rcept_no,
        "is_amendment": r.is_amendment,
        "is_foreign": r.is_foreign,
        "raw_tables_count": r.raw_tables_count,
        "notes": list(r.notes),
    }


def _dict_to_record(d: dict) -> TrancheRecord:
    """meta.json dict → TrancheRecord 복원."""
    def parse_date(v):
        if isinstance(v, str) and len(v) == 10 and v[4] == "-":
            try:
                return date.fromisoformat(v)
            except ValueError:
                return v
        return v
    return TrancheRecord(
        subscription_date=parse_date(d.get("subscription_date")),
        issuer_alias=d.get("issuer_alias", ""),
        issuer_full=d.get("issuer_full", ""),
        corp_code=d.get("corp_code", ""),
        series=d.get("series", ""),
        bond_type=d.get("bond_type", ""),
        credit_rating=d.get("credit_rating", ""),
        maturity=parse_date(d.get("maturity")),
        initial_amount=d.get("initial_amount"),
        issue_limit=d.get("issue_limit"),
        demand_amount=d.get("demand_amount"),
        final_amount=d.get("final_amount"),
        series_total=d.get("series_total"),
        rate_target=d.get("rate_target", ""),
        rate_demand=d.get("rate_demand", ""),
        rate_final=d.get("rate_final"),
        lead_managers=list(d.get("lead_managers", [])),
        underwriter_alloc=dict(d.get("underwriter_alloc", {})),
        rcept_no=d.get("rcept_no", ""),
        is_amendment=d.get("is_amendment", False),
        is_foreign=d.get("is_foreign", False),
        raw_tables_count=d.get("raw_tables_count", 0),
        notes=list(d.get("notes", [])),
    )


def save_meta(output_path: Path, records: list[TrancheRecord],
              processed_rcept_nos: list[str] | None = None) -> Path:
    """엑셀 옆에 사이드카 .meta.json 저장 — incremental update 용.

    records 직렬화 + 처리된 rcept_no list 보관. update 시 이 메타로 기존 records 복원
    + 새 fetch 에서 이미 처리된 rcept_no 건너뛰기.
    """
    meta_path = output_path.with_suffix(".meta.json")
    rcepts = set(processed_rcept_nos or [])
    rcepts.update(r.rcept_no for r in records if r.rcept_no)
    meta = {
        "records": [_record_to_dict(r) for r in records],
        "processed_rcept_nos": sorted(rcepts),
    }
    meta_path.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return meta_path


def load_meta(xlsx_path: Path) -> tuple[list[TrancheRecord], set[str]]:
    """엑셀 옆 .meta.json 로드 → (records, processed_rcept_no set)."""
    meta_path = xlsx_path.with_suffix(".meta.json")
    if not meta_path.exists():
        return [], set()
    data = json.loads(meta_path.read_text(encoding="utf-8"))
    records = [_dict_to_record(d) for d in data.get("records", [])]
    rcepts = set(data.get("processed_rcept_nos", []))
    return records, rcepts


def read_existing_keys(xlsx_path: Path) -> set[tuple[str, str]]:
    """xlsx 의 모든 시트에서 데이터 행의 (발행사, 회차) 키 셋 추출.

    update 시작 전 호출해서 xlsx 의 실제 행 상태를 파악 (사용자가 수동으로
    행을 추가/삭제한 결과를 반영). update 후 변화량 계산용.
    .meta.json 의 records 와 별개 — xlsx 의 실제 행만 본다.
    """
    keys: set[tuple[str, str]] = set()
    if not xlsx_path.exists():
        return keys
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return keys
    issuer_col = config.COL["발행사"]
    series_col = config.COL["회차"]
    max_col = max(issuer_col, series_col)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=3, max_col=max_col, values_only=True):
            issuer = row[issuer_col - 1] if len(row) >= issuer_col else None
            series = row[series_col - 1] if len(row) >= series_col else None
            if issuer:
                keys.add((str(issuer), str(series) if series is not None else ""))
    wb.close()
    return keys


def backup_current_state(output_path: Path, max_keep: int = 5) -> Path | None:
    """write() 직전 현재 xlsx + meta.json 을 .bak 폴더에 timestamped snapshot 으로 보관.

    write() 가 _build_template 으로 xlsx 를 destruct + rebuild 하기 전에 호출 →
    write 도중 에러로 xlsx 가 빈 상태로 저장되거나 의도치 않은 데이터 손실 시 즉시
    복구 가능.

    구조: <xlsx_dir>/.bak/<YYYYMMDD-HHMMSS>/{DCM Table.xlsx, DCM Table.meta.json}
    rolling: 가장 최신 max_keep 개 snapshot 만 유지, 나머지 삭제.
    """
    if not output_path.exists():
        return None  # 첫 생성 케이스 — 백업할 게 없음
    try:
        meta_path = output_path.with_suffix(".meta.json")
        bak_root = output_path.parent / ".bak"
        bak_root.mkdir(exist_ok=True)

        from datetime import datetime as _dt
        ts = _dt.now().strftime("%Y%m%d-%H%M%S")
        snapshot = bak_root / ts
        snapshot.mkdir(exist_ok=True)

        shutil.copy2(output_path, snapshot / output_path.name)
        if meta_path.exists():
            shutil.copy2(meta_path, snapshot / meta_path.name)

        # 오래된 snapshot 정리 — 이름이 timestamp 라 자연 정렬됨
        snapshots = sorted([d for d in bak_root.iterdir() if d.is_dir()],
                           key=lambda d: d.name, reverse=True)
        for old in snapshots[max_keep:]:
            shutil.rmtree(old, ignore_errors=True)
        return snapshot
    except Exception:
        # 백업 실패해도 main write 는 계속 — 백업은 보조 안전망
        return None


def _copy_header_template(dst_ws, src_ws=None) -> None:
    """src_ws (또는 template) 에서 헤더 + 스타일 + 필터 + 너비 복사.

    신규 연도 시트 생성 시 사용. src_ws 가 None 이면 template_dcm_sheet.xlsx 의
    2024년 시트 사용. 현재 wb 의 기존 연도 시트가 있으면 그걸 전달하면 완전히
    동일한 형태 유지.
    포함:
      - 헤더 cells (1-2행) + 스타일
      - 컬럼 너비 (모든 columns)
      - 헤더 영역 (1-2행) merged_cells
      - 행 높이 (1-2)
      - auto_filter (헤더 행 기준, 데이터 채울 때 ref 자동 갱신)
      - freeze_panes (A3)
    """
    from copy import copy as _copy
    from openpyxl.worksheet.views import Selection

    src_wb = None
    if src_ws is None:
        src_wb = load_workbook(config.TEMPLATE_XLSX)
        src_ws = src_wb["2024년"] if "2024년" in src_wb.sheetnames else src_wb[src_wb.sheetnames[0]]

    # 헤더 cells (1-2 row)
    for row_num in (1, 2):
        for cell in src_ws[row_num]:
            new_cell = dst_ws.cell(row=row_num, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = _copy(cell.font)
                new_cell.fill = _copy(cell.fill)
                new_cell.border = _copy(cell.border)
                new_cell.alignment = _copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = _copy(cell.protection)

    # 컬럼 너비 (A~BT)
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = dim.width

    # 행 높이 (1-2 헤더만)
    for row_num in (1, 2):
        if row_num in src_ws.row_dimensions:
            sd = src_ws.row_dimensions[row_num]
            if sd.height is not None:
                dst_ws.row_dimensions[row_num].height = sd.height

    # 헤더 영역 (1-2행) merged_cells
    for mc in src_ws.merged_cells.ranges:
        if mc.max_row <= 2:
            dst_ws.merge_cells(str(mc))

    # auto_filter — 헤더 기준 (데이터 row 추가 시 caller 가 ref 갱신 권장)
    if src_ws.auto_filter.ref:
        # 헤더 행 기준만 우선 (예: A2:BT2). 데이터 채워지면 갱신.
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(src_ws.max_column)
        dst_ws.auto_filter.ref = f"A2:{last_col_letter}2"

    # freeze panes
    dst_ws.freeze_panes = "A3"
    dst_ws.sheet_view.topLeftCell = "A1"
    if dst_ws.sheet_view.pane is not None:
        dst_ws.sheet_view.pane.topLeftCell = "A3"
    dst_ws.sheet_view.selection = [Selection(activeCell="A3", sqref="A3")]

    # 화면 배율 — 다른 연도 시트와 통일 (기본 85%). src_ws zoom 이 설정돼 있으면 그걸 사용.
    src_zoom = getattr(src_ws.sheet_view, "zoomScale", None)
    dst_ws.sheet_view.zoomScale = src_zoom if src_zoom else 85
    dst_ws.sheet_view.zoomScaleNormal = src_zoom if src_zoom else 85

    # 표준 컬럼 너비 강제 적용 — 모든 시트 통일 위해 src 값 무시하고 덮어씀.
    # 금액 5 cols (G-K): 5.5 (73px). 브로커 cols (주관 P 부터 인수 끝까지): 4.42 (60px).
    _enforce_standard_widths(dst_ws)


def _enforce_standard_widths(ws) -> None:
    """표준 컬럼 너비 적용 — 신규 시트 생성 & broker 컬럼 확장 시 호출.

    주의: openpyxl 의 width 값 (XML 저장값) 과 Excel UI 표시 width 가 다름.
    Excel UI 표시 = XML width − 0.58 (Calibri 11 / 맑은 고딕 11 기준).
    사용자 목표:
      - 금액 5 cols (최초모집/발행한도/수요예측/최종발행/회차합산):
        Excel UI 5.50 (73 px) → XML width = 6.08
      - 주관/인수 broker 컬럼: Excel UI 4.42 (60 px) → XML width = 5.00
    """
    from openpyxl.utils import get_column_letter
    # 금액 5 cols — XML 6.08 → Excel UI 5.50
    for letter in ("G", "H", "I", "J", "K"):
        ws.column_dimensions[letter].width = 6.08203125
    # 브로커 영역 — 주관/인수 merged_cells 범위 자동 감지. XML 5.0 → Excel UI 4.42
    lead_mc, uw_mc = _find_group_merges(ws)
    if lead_mc and uw_mc:
        start_col = lead_mc[0]
        end_col = uw_mc[1]
        for c in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 5.0


def write(records: list[TrancheRecord], output_path: Path,
          processed_rcept_nos: list[str] | None = None) -> None:
    """records를 청약일 연도별로 분리해서 시트에 기록 + meta.json 저장."""
    # write 직전 현재 상태 백업 (rolling 5). _build_template 의 destruct + rebuild
    # 도중 에러 발생 시 .bak 에서 즉시 복구 가능.
    backup_current_state(output_path)
    _build_template(output_path)
    wb = load_workbook(output_path)

    # 연도별로 분리
    by_year: dict[str, list[TrancheRecord]] = {}
    for r in records:
        if isinstance(r.subscription_date, (date, datetime)):
            year = str(r.subscription_date.year)
        else:
            year = "기타"
        key = f"{year}년"
        by_year.setdefault(key, []).append(r)

    # 청약일 → DART 공시 접수순(rcept_no asc, 즉 먼저 접수된 것 먼저) → 회차 순 정렬.
    # 발행사명(한/영) 가나다 정렬은 청약일 같은 그룹 안에서 의미 없고 오히려 공시순과 어긋남.
    for sheet_name in by_year:
        by_year[sheet_name].sort(key=lambda x: (
            x.subscription_date or date(1900, 1, 1),
            x.rcept_no,
            x.series,
        ))

    for sheet_name, recs in by_year.items():
        if sheet_name not in wb.sheetnames:
            # 신규 연도 시트: 현재 wb 의 기존 연도 시트를 src 로 우선 사용 (완전 일치 보장).
            # 없으면 _copy_header_template 가 template_dcm_sheet.xlsx 에서 복사.
            template_src = None
            for candidate in ("2024년", "2025년", "2026년"):
                if candidate in wb.sheetnames and candidate != sheet_name:
                    template_src = wb[candidate]
                    break
            new_ws = wb.create_sheet(sheet_name)
            _copy_header_template(new_ws, template_src)
        ws = wb[sheet_name]
        # 시작 행: 헤더 다음
        start_row = 3
        for i, rec in enumerate(recs):
            _write_tranche_row(ws, start_row + i, rec)

        # 다중트랜치 발행건의 H/K 셀 병합
        _merge_multi_tranche(ws, recs, start_row)

        # 데이터 행까지 포함해 다시 한번 셀 테두리 정규화
        _normalize_borders(ws)

        # 표준 width 강제 — 매 write 마다 모든 시트에 재적용해서 통일 보장.
        # (cmd_update 에서 신규 broker 컬럼 추가됐을 때 기존 시트의 KR 등 기존 broker
        # 컬럼 width 가 default 8 로 남는 문제 해결.)
        _enforce_standard_widths(ws)

        # zoom 도 매번 통일 (85)
        ws.sheet_view.zoomScale = 85
        ws.sheet_view.zoomScaleNormal = 85

        # auto_filter ref 갱신 (데이터 마지막 행까지)
        if recs and ws.auto_filter.ref:
            from openpyxl.utils import get_column_letter
            last_row = start_row + len(recs) - 1
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

    # 시트 정렬: 연도 시트 내림차순 (2026 → 2025 → 2024 → 2023 ...) + '기타' 마지막
    year_sheets = sorted(
        [s for s in wb.sheetnames if s.endswith("년") and s[:-1].isdigit()],
        key=lambda s: int(s[:-1]), reverse=True,
    )
    other_sheets = [s for s in wb.sheetnames if s not in year_sheets]
    desired_order = year_sheets + other_sheets
    # openpyxl 의 sheet 정렬: _sheets 직접 재배치
    wb._sheets = [wb[name] for name in desired_order]

    wb.save(output_path)

    # 저장 후 XML 직접 수정: '텍스트로 저장된 숫자' 오류 표시 숨김 (회차 컬럼)
    _add_ignored_errors_xml(output_path)

    # 사이드카 메타 저장 (incremental update 용 records + rcept 누적)
    save_meta(output_path, records, processed_rcept_nos)


def _merge_multi_tranche(ws, recs: list[TrancheRecord], start_row: int):
    """동일 발행건의 트랜치 행 H, K 열 병합 (회차 prefix '74-1','74-2'식 매칭)."""
    if not recs:
        return

    def base_series(s: str) -> str:
        # "74-1" → "74", "74-2" → "74", "74" → "74"
        return s.split("-")[0] if "-" in s else s

    i = 0
    while i < len(recs):
        j = i
        cur_issuer = recs[i].issuer_alias
        cur_base = base_series(recs[i].series)
        while (j + 1 < len(recs)
               and recs[j + 1].issuer_alias == cur_issuer
               and base_series(recs[j + 1].series) == cur_base):
            j += 1
        if j > i:
            row_top = start_row + i
            row_bot = start_row + j
            ws.merge_cells(start_row=row_top, end_row=row_bot,
                            start_column=config.COL["발행한도"], end_column=config.COL["발행한도"])
            ws.merge_cells(start_row=row_top, end_row=row_bot,
                            start_column=config.COL["회차합산"], end_column=config.COL["회차합산"])
        i = j + 1
