"""DCM 시트 셀 수식 생성 (공통 모듈)."""
from __future__ import annotations
from openpyxl.utils import get_column_letter
import config


def lead_to_uw_col_letter(alias: str) -> str | None:
    """주관 약칭 → 같은 회사의 인수 컬럼 letter (예: 'KB' → 'AO')."""
    if alias not in config.UNDERWRITERS:
        return None
    idx = config.UNDERWRITERS.index(alias)  # 0-based
    return get_column_letter(config.underwriter_col_start() + idx)


def build_lead_formula(row: int, my_alias: str, all_lead_aliases: list[str]) -> str:
    """주관사 셀에 들어갈 수식.

    산식: 본인 인수금액 + (전체 인수합 - 모든 주관사 본인 인수합) / 주관사 수
    - 주관사 1곳: =SUM(AOr:BTr)  (전체 인수금액을 그 곳이 받음)
    - 주관사 N곳: =MyCol_r + (SUM(AOr:BTr) - (Lead1Col_r + Lead2Col_r + ...)) / N
    """
    n = len(all_lead_aliases)
    if n == 0:
        return ""

    uw_start = config.underwriter_col_start()
    uw_end = uw_start + len(config.UNDERWRITERS) - 1
    ao = get_column_letter(uw_start)
    bt = get_column_letter(uw_end)
    total = f"SUM({ao}{row}:{bt}{row})"
    if n == 1:
        return f"={total}"

    lead_uw_cols = [lead_to_uw_col_letter(a) for a in all_lead_aliases]
    lead_uw_cols = [c for c in lead_uw_cols if c]
    if not lead_uw_cols:
        return ""
    leads_sum = "+".join(f"{c}{row}" for c in lead_uw_cols)

    my_col = lead_to_uw_col_letter(my_alias)
    if my_col:
        return f"={my_col}{row}+({total}-({leads_sum}))/{n}"
    # 주관이지만 인수에는 안 잡힌 케이스 (드물지만 안전망)
    return f"=({total}-({leads_sum}))/{n}"
