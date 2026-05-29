"""ECM (지분증권) 데이터 수집 진입점 — IPO + 유상증자.

사용:
  py auto/main_ecm.py collect 2026-01-01 2026-05-22
  py auto/main_ecm.py collect 2026-01-01 2026-05-22 --corp-code 00126380
  py auto/main_ecm.py status

흐름:
  1. dart_client.list_ecm_filings (C001+C003) → 기간 내 ECM 공시 목록
  2. classify_filing 으로 종류 분류 (stage1 / amend / final / report / ignore)
  3. corp_code 별 그룹 + 시간순 정렬 → 같은 회사의 공시들을 묶음
  4. 한 회사의 공시 묶음에서 stage1 등장 시점마다 새 "딜" 시작 (= 다음 stage1 전까지가 한 딜)
  5. 각 딜에 대해:
     a) stage1 본문에서 parse_offering_summary → IPO/유상증자 판별
     b) 분류에 맞춰 parser 호출, amend 들로 일정 갱신
     c) IPO: KIND 에서 상장예정일 조회
  6. 시간순 정렬 후 ECM Table test.xlsx 의 IPO/유상증자 시트에 쓰기
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

# auto/ 폴더가 sys.path 에 있어야 sibling import 동작
sys.path.insert(0, str(Path(__file__).resolve().parent))

import json
import openpyxl

import dart_client
import kind_client
import parser_ecm
import config_ecm
import config  # mappings.json 경로용 (DCM 와 공유)


# ============== 딜 묶음 매칭 ==============

@dataclass
class DealGroup:
    """한 회사의 한 ECM 발행 딜에 속한 공시들 (시간순)."""
    corp_code: str
    corp_name: str
    stage1: Optional["dart_client.Filing"] = None
    amends: list = field(default_factory=list)
    finals: list = field(default_factory=list)   # IPO=1개 / 유상=1·2차
    reports: list = field(default_factory=list)  # IPO 의 증권발행실적보고서
    is_withdrawn: bool = False  # 철회된 딜 (사용자 룰 2026-05-26: 수집 대상 제외)


def group_into_deals(filings: list) -> list[DealGroup]:
    """필링 리스트를 corp_code 별로 묶고, stage1 등장 시점마다 새 딜 시작.

    같은 corp_code 의 공시들이 시간 순서대로 들어옴:
      [stage1_A, amend, amend, final, report, stage1_B, amend, final, ...]
      → 두 딜로 분리: [A: stage1+amend+amend+final+report], [B: stage1+amend+final]

    "stage1_backfill" 처리:
      - 진짜 stage1 이 우리 수집 기간 밖일 경우, [정정제출요구]/[첨부추가] 공시의
        본문에 stage1 정보가 그대로 들어있음 → 첫 backfill 을 임시 stage1 으로 승격
      - 이미 진행 중인 딜이 있다면 같은 딜의 amend 로 합침 (배정기준일/납입일 변동 체크용)

    corp_code 없는 공시 (드물게 발생) 는 corp_name 으로 fallback.
    """
    # corp_code 별 그룹
    by_corp: dict[str, list] = {}
    for f in sorted(filings, key=lambda x: x.rcept_no):
        key = f.corp_code or f.corp_name
        by_corp.setdefault(key, []).append(f)

    deals: list[DealGroup] = []
    for key, fs in by_corp.items():
        current: Optional[DealGroup] = None
        for f in fs:
            kind = parser_ecm.classify_filing(f.report_nm)
            if kind == "stage1":
                # 새 딜 시작 (직전 딜 flush)
                if current is not None:
                    deals.append(current)
                current = DealGroup(
                    corp_code=f.corp_code, corp_name=f.corp_name, stage1=f,
                )
            elif kind == "stage1_backfill":
                if current is None or current.is_withdrawn:
                    # 진행 중인 딜 없음 또는 withdrawn → 새 deal 로 시작
                    # (사용자 룰 2026-05-26: 에스투더블유 — 철회 후 별개 IPO 시작)
                    if current is not None:
                        deals.append(current)
                    current = DealGroup(
                        corp_code=f.corp_code, corp_name=f.corp_name, stage1=f,
                    )
                else:
                    # 이미 stage1 이 잡힌 상태 (정상) → amend 처럼 합침
                    current.amends.append(f)
            elif kind == "amend":
                if current is None:
                    # stage1 없는 상태에서 amend → 임시 deal 시작 (stage1=None, 나중에 dropdown 으로 백필)
                    current = DealGroup(
                        corp_code=f.corp_code, corp_name=f.corp_name, stage1=None,
                    )
                current.amends.append(f)
            elif kind == "final":
                if current is None:
                    current = DealGroup(
                        corp_code=f.corp_code, corp_name=f.corp_name, stage1=None,
                    )
                current.finals.append(f)
            elif kind == "report":
                if current is not None:
                    # 이미 진행 중인 deal 이 있을 때만 reports 에 합침
                    current.reports.append(f)
                # else: report-only corp 은 ECM 활성 corp 으로 추정하기 어려움
                # ("증권발행실적보고서" report_nm 필터가 채권 발행실적도 잡으므로
                #  채권사가 자기 채권 발행실적을 올린 corp 가 다수 포함됨 — 무시)
            elif kind == "withdrawn":
                # 사용자 룰 2026-05-26: 철회신고서 → 현재 진행 중 deal 만 철회 처리.
                # (다음 stage1 등장 시점에 새 deal 시작 → 그건 별개 정상 deal)
                if current is not None:
                    current.is_withdrawn = True
            # rm="철" 마크 (rm 비고 필드) 도 보조 신호 — 같은 deal 의 어떤 Filing 에든
            # "철" 마크 있으면 withdrawn 처리
            if current is not None and "철" in (getattr(f, "rm", "") or ""):
                current.is_withdrawn = True
            # ignore 는 skip
        if current is not None:
            deals.append(current)
    return deals


# ============== 한 딜 처리 ==============

@dataclass
class DealResult:
    """한 딜의 처리 결과 — IPO 또는 유상증자 record."""
    kind: str = ""  # "ipo" / "rights"
    ipo_record: Optional[parser_ecm.IPORecord] = None
    rights_record: Optional[parser_ecm.RightsRecord] = None
    # 시간순 정렬 키 (IPO=상장일, 유상=배정기준일). 없으면 stage1 rcept_dt
    sort_key: Optional[date] = None
    # 보조 데이터 (write_to_xlsx 가 사용)
    listing_date_planned: Optional[date] = None
    underwriter_amounts_eok: dict[str, float] = field(default_factory=dict)
    lead_aliases: set = field(default_factory=set)
    lead_perf: dict[str, float] = field(default_factory=dict)


def process_deal(deal: DealGroup) -> Optional[DealResult]:
    """한 딜 처리 — stage1 본문에서 IPO/유상증자 판별 후 모든 단계 처리."""
    if deal.stage1 is None:
        return None

    # stage1 본문 fetch (제2부 컷오프)
    secs_stage1 = dart_client.fetch_ecm_stage1_document(deal.stage1.rcept_no)
    summ = parser_ecm.parse_offering_summary(secs_stage1)
    kind = summ.get("kind", "unknown")
    if kind not in ("ipo", "rights"):
        print(f"  [WARN] {deal.corp_name}: 분류 실패 — {summ}")
        return None

    res = DealResult(kind=kind)
    if kind == "ipo":
        res.ipo_record = _process_ipo(deal, secs_stage1)
        # 상장(예정)일 = 향후 report 의 확정 상장일 우선, 없으면 KIND 예정일
        res.sort_key = res.listing_date_planned
    else:
        res.rights_record = _process_rights(deal, secs_stage1)
        res.sort_key = res.rights_record.record_date if res.rights_record else None
    return res


def _process_ipo(deal: DealGroup, secs_stage1: dict) -> parser_ecm.IPORecord:
    """IPO 딜 처리: stage1 + final + report + KIND 통합."""
    rec = parser_ecm.parse_ipo_stage1(
        secs_stage1, rcept_no=deal.stage1.rcept_no,
        corp_name=deal.corp_name, corp_code=deal.corp_code,
    )

    # final ([발행조건확정]) — 사용자 룰 2026-05-25: IPO 첫 번째 [발행조건확정] 만 보고,
    # 거기서 인수단 정정 후 표 → broker별 amount_won 직접 추출. 보고서는 broker 정보용 아님.
    if deal.finals:
        final_f = deal.finals[0]  # **첫 번째** [발행조건확정] (가격·broker 확정 시점)
        rec.rcept_no_final = final_f.rcept_no
        secs_f = dart_client.fetch_full_document(
            final_f.rcept_no, title_predicate=dart_client.ecm_final_strict_predicate)
        final = parser_ecm.parse_ipo_final(secs_f, final_f.rcept_no)
        if final.get("final_qty"):
            rec.final_qty = final["final_qty"]
        if final.get("final_price"):
            rec.final_price = final["final_price"]
        uw_rows = final.get("underwriter_rows", []) or []

        # stage1 fallback — final 본문에 인수단 표 없으면 (단일 broker IPO 등 단순 케이스)
        # stage1 본문에서 broker 비율을 가져와서 정정 후 final_price 로 amount 재계산.
        # 사용자 룰 2026-05-25: "다수 broker 라도 final 에 표 없으면 stage1 비율 유지 +
        #                       정정 후 가액 기준 총액 비례 분배"
        if not uw_rows and deal.stage1:
            try:
                import pandas as _pd
                import io as _io
                secs_s1 = dart_client.fetch_full_document(
                    deal.stage1.rcept_no,
                    title_predicate=dart_client.ecm_stage1_strict_predicate)
                stage1_uw = []
                for html in secs_s1.values():
                    try:
                        tables = _pd.read_html(_io.StringIO(html), flavor="lxml")
                    except Exception:
                        continue
                    for t in tables:
                        t = parser_ecm._promote_header(t)
                        cols = [str(c) for c in t.columns]
                        cols_joined = " ".join(cols)
                        if not (("인수인" in cols_joined or "인수(주선)인" in cols_joined)
                                and "인수수량" in cols_joined):
                            continue
                        role_col = name_col = qty_col = None
                        for ci, c in enumerate(cols):
                            cs = str(c).replace(" ", "")
                            if cs in ("인수인", "인수(주선)인") and role_col is None:
                                role_col = ci
                            elif (cs.startswith("인수인") or cs.startswith("인수(주선)인")) \
                                    and name_col is None and ci != role_col:
                                name_col = ci
                            elif "인수수량" in cs:
                                qty_col = ci
                        if role_col is None or name_col is None or qty_col is None:
                            continue
                        for _, row in t.iterrows():
                            rl = str(row.iloc[role_col]).replace(" ", "")
                            nm = str(row.iloc[name_col]).replace(" ", "")
                            try:
                                qty_int = int(str(row.iloc[qty_col]).replace(",", ""))
                            except (ValueError, TypeError):
                                continue
                            if not nm or nm in ("nan", "None", "계", "합계"):
                                continue
                            stage1_uw.append({"role": rl, "name": nm, "qty": qty_int})
                        if stage1_uw:
                            break
                    if stage1_uw:
                        break

                # broker qty 비례 + final_qty / final_price 로 amount 재계산
                if stage1_uw:
                    s_total = sum(x["qty"] for x in stage1_uw) or 1
                    fp = rec.final_price
                    fq = rec.final_qty
                    for x in stage1_uw:
                        # final_qty 와 stage1 qty_sum 다르면 비례 조정
                        if fq and fq != s_total:
                            new_qty = round(x["qty"] * fq / s_total)
                        else:
                            new_qty = x["qty"]
                        x["qty"] = new_qty
                        x["amount_won"] = (new_qty * fp) if fp else None
                    uw_rows = stage1_uw
                    print(f"  [INFO] {deal.corp_name}: final 본문에 인수단 표 없음 "
                          f"→ stage1 fallback ({len(stage1_uw)} broker, "
                          f"비율 유지 + 정정 후 가액 재계산)")
            except Exception as e:
                print(f"  [WARN] {deal.corp_name}: stage1 fallback 실패 — {e}")

        # final 의 인수단 (수량 + 금액) — broker alias 매핑 위해 보관
        rec._final_underwriter_rows = uw_rows
        # 기관 청약수량 = final 의 수요예측 합계
        if final.get("inst_subscribed_demand") is not None:
            rec.inst_subscribed = final["inst_subscribed_demand"]

        # [정정]투자설명서 fallback — 공시 기업이 [발행조건확정] 본문에 수요예측 결과
        # 표를 누락한 케이스 (엔에이치기업인수목적25호 2022-10 케이스). final 직후
        # 1개월 안에 [정정]투자설명서가 별도로 올라와 누락 내용을 채워넣음.
        # 사용자 룰 2026-05-26: [발행조건확정] 본문에서 못 잡은 경우에만 추가 fetch
        # — 정상 케이스에서는 DART 호출 0 (불필요 부하 방지).
        if rec.inst_subscribed is None and deal.corp_code and deal.finals:
            from datetime import date as _dt_date, timedelta as _td
            f_rcept = deal.finals[-1].rcept_no
            try:
                d_start = _dt_date(int(f_rcept[:4]), int(f_rcept[4:6]), int(f_rcept[6:8]))
                d_end = d_start + _td(days=30)
                extras = []
                for cat in ("C001", "C003"):
                    extras += dart_client.list_filings(
                        d_start, d_end, only_bond_registration=False,
                        corp_code=deal.corp_code, pblntf_detail_ty=cat)
                seen = set()
                for ef in extras:
                    if ef.rcept_no in seen:
                        continue
                    seen.add(ef.rcept_no)
                    rn = (ef.report_nm or "").replace(" ", "")
                    if "투자설명서" not in rn:
                        continue
                    try:
                        extra_html = dart_client.fetch_full_document(ef.rcept_no)
                        extra_demand = parser_ecm.extract_inst_demand_qty(extra_html)
                        if extra_demand:
                            rec.inst_subscribed = extra_demand
                            print(f"  [INFO] {deal.corp_name}: inst_subscribed "
                                  f"{extra_demand:,} ([정정]투자설명서 "
                                  f"{ef.rcept_no} fallback)")
                            break
                    except Exception:
                        continue
            except Exception as e:
                print(f"  [WARN] {deal.corp_name}: 투자설명서 fallback 실패 — {e}")

    # report (증권발행실적보고서)
    # 사용자 룰 2026-05-25: deal.reports 가 비어있는데 final 후 시간 지나면 corp_code
    # 검색으로 직접 fetch (cmd_collect 시점 group_into_deals 가 못 잡은 케이스 보완).
    # 케이뱅크/에스팀/액스비스 등 final 후 며칠 만에 보고서 올라온 케이스.
    if not deal.reports and rec.rcept_no_final and deal.corp_code:
        try:
            from datetime import date as _date, timedelta as _td
            f_rcept = rec.rcept_no_final
            d_start = _date(int(f_rcept[:4]), int(f_rcept[4:6]), int(f_rcept[6:8]))
            d_end = _date.today()  # final ~ 오늘 사이 검색
            if d_end > d_start:
                filings = dart_client.list_ecm_filings(
                    start=d_start, end=min(d_end, d_start + _td(days=90)),
                    corp_code=deal.corp_code)
                for f in filings:
                    if "증권발행실적보고서" in (f.report_nm or "").replace(" ", ""):
                        f.corp_name = deal.corp_name
                        f.corp_code = deal.corp_code
                        deal.reports.append(f)
                        print(f"  [+report] {deal.corp_name}: corp_code 검색으로 "
                              f"보고서 {f.rcept_no} 추가")
                        break
        except Exception as e:
            print(f"  [WARN] {deal.corp_name}: corp_code 보고서 검색 실패 — {e}")

    report_listing_date = None
    if deal.reports:
        report_f = deal.reports[-1]
        secs_r = dart_client.fetch_full_document(
            report_f.rcept_no, title_predicate=dart_client.ecm_report_strict_predicate)
        report = parser_ecm.parse_ipo_report(secs_r, report_f.rcept_no)
        for k in ["esop_initial", "esop_final",
                  "inst_initial", "inst_final",
                  "general_initial", "general_subscribed", "general_final"]:
            if report.get(k) is not None:
                setattr(rec, k, report[k])
        # 기관 청약수량 (M 컬럼) — 사용자 룰 2026-05-25 명시:
        # **첫 번째 [발행조건확정] 공시의 수요예측 합** 에서만 채움.
        # 보고서의 청약현황수량은 의미가 다름 (M 으로 사용 금지).
        # → 보고서 fallback 폐기 (rec.inst_subscribed 가 None 이면 그대로 None 유지)
        rec._report_underwriter_amounts = report.get("underwriter_amounts", [])
        # 확정 상장일 (사용자 룰 — IV. 증권교부일 등의 3. 상장일(매매개시일))
        report_listing_date = report.get("listing_date_confirmed")
        # 연도 오기 자동 보정 — 신성에스티 케이스 (보고서에 "2022.10.19" 단순 오기,
        # 실제 2023-10-19). 상장일 < stage1 공시일이면 연도+1 보정.
        if report_listing_date and rec.rcept_no_stage1:
            report_listing_date = parser_ecm._maybe_fix_listing_date_typo(
                report_listing_date, rec.rcept_no_stage1)

    # 상장일 결정: report 의 확정 상장일이 1순위, KIND 의 예정일이 2순위.
    # KIND 도 못 찾으면 None → write 단계에서 "미정" 표기.
    # 추가 안전망: report 가 stage1 보다 과거이면 KIND 우선 (연도+1 보정으로도
    # 안 잡힌 케이스 대비)
    use_kind = False
    if not report_listing_date:
        use_kind = True
    elif rec.rcept_no_stage1 and len(rec.rcept_no_stage1) >= 8:
        try:
            from datetime import date as _d
            stage1_date = _d(int(rec.rcept_no_stage1[:4]),
                             int(rec.rcept_no_stage1[4:6]),
                             int(rec.rcept_no_stage1[6:8]))
            if report_listing_date < stage1_date:
                print(f"  [WARN] {deal.corp_name}: 보고서 상장일 {report_listing_date} "
                      f"< stage1 공시일 {stage1_date} — KIND 조회로 대체 시도")
                use_kind = True
        except ValueError:
            pass

    if not use_kind:
        rec._listing_date = report_listing_date
    else:
        try:
            sched = kind_client.fetch_listing_schedule(deal.corp_name)
            if sched and sched.listing_date_planned:
                rec._listing_date = sched.listing_date_planned
            elif report_listing_date:
                # KIND 실패 + report 비정상 → 그래도 report 값 유지 (None 보다 낫음)
                rec._listing_date = report_listing_date
        except Exception as e:
            print(f"  [WARN] {deal.corp_name}: KIND 조회 실패 — {e}")
            rec._listing_date = report_listing_date  # fallback

    return rec


def _process_rights(deal: DealGroup, secs_stage1: dict) -> parser_ecm.RightsRecord:
    """유상증자 딜 처리: stage1 + amend 일정 갱신 + final ×1~2."""
    rec = parser_ecm.parse_rights_stage1(
        secs_stage1, rcept_no=deal.stage1.rcept_no,
        corp_name=deal.corp_name, corp_code=deal.corp_code,
    )

    # stage1 "2. 공모방법" 에 발행주식총수 없으면 → "4. 주식의 총수 등" fetch.
    # 사용자 룰 2026-05-25 명시: offering_type 무관 — stage1 에서 못 찾으면 무조건 봐야.
    # (이전 코드는 "일반공모/제3자배정" 한정이라 뉴로메카 "주주배정후 실권주 일반공모"
    #  같은 케이스가 누락됐음.)
    if rec.existing_qty is None:
        try:
            secs_ts = dart_client.fetch_full_document(
                deal.stage1.rcept_no,
                title_predicate=dart_client.ecm_total_shares_predicate)
            if secs_ts:
                eq = parser_ecm.extract_existing_qty_from_total_shares(secs_ts)
                if eq is not None:
                    rec.existing_qty = eq
        except Exception as e:
            print(f"  [WARN] {deal.corp_name} 4. 주식의 총수 등 fallback "
                  f"fetch 실패: {e}")

    # 3차 fallback (옵션 A — 2026-05-25): "4. 주식의 총수 등" 본문이 회사 측 생략된
    # 케이스 (아미코젠 2025-12) — III. 투자위험요소 sub-section 만 추가 fetch.
    # parse_rights_stage1 의 패턴 3 ("현재 발행주식 총수") / 패턴 4 ("발행주식 총수인")
    # 가 III 본문 매칭. 분량 부담 줄이려 sub-section (사업/회사/기타위험) 만 fetch.
    if rec.existing_qty is None:
        try:
            secs_risk = dart_client.fetch_full_document(
                deal.stage1.rcept_no,
                title_predicate=dart_client.ecm_risk_factors_predicate)
            if secs_risk:
                rec2 = parser_ecm.parse_rights_stage1(
                    secs_risk, rcept_no=deal.stage1.rcept_no,
                    corp_name=deal.corp_name, corp_code=deal.corp_code)
                if rec2.existing_qty is not None:
                    rec.existing_qty = rec2.existing_qty
                    print(f"  [INFO] {deal.corp_name}: existing_qty "
                          f"{rec.existing_qty:,} (III. 투자위험요소 sub-section "
                          f"fallback — 4. 주식의 총수 등 본문 생략 케이스)")
        except Exception as e:
            print(f"  [WARN] {deal.corp_name} III. 투자위험요소 fallback "
                  f"fetch 실패: {e}")

    # 제3자배정 fallback: stage1 본문에 이사회 결의일 텍스트 없음 → 주요사항보고서
    # (유상증자결정) 본문 fetch → 이사회결의일(결정일) 행 추출 (윙입푸드 2024-03 케이스).
    # **lazy** — 정말 stage1 에 없을 때만. 부하 최소화.
    if rec.record_date is None and rec.offering_type == "제3자배정":
        try:
            msr_rcept = parser_ecm.extract_msr_rcept_no(secs_stage1)
            if msr_rcept:
                msr_secs = dart_client.fetch_full_document(msr_rcept)
                bd = parser_ecm.extract_board_resolution_date_from_msr(msr_secs)
                if bd is not None:
                    rec.record_date = bd
                    print(f"  [INFO] {deal.corp_name}: 이사회 결의일 {bd} "
                          f"(주요사항보고서 {msr_rcept} 에서 추출)")
        except Exception as e:
            print(f"  [WARN] {deal.corp_name} 주요사항보고서 fallback 실패: {e}")

    # 일정 (배정기준일/납입일) 갱신은 rcept_no 우선순위 기반 — 가장 늦은 공시의
    # non-None 값만 채택. 회사가 stage1 에 오기한 일정이 amend 에서 정정됐으나
    # 그 사이 final 본문이 stage1 오기를 복제한 경우 (예: 대성창투 2022→2023 정정)
    # 시간순으로 처리하면 amend 갱신이 final 처리에서 무효화되는 버그 방지.
    latest_rd_rcept = deal.stage1.rcept_no  # stage1 의 record_date 가 첫 값
    latest_pd_rcept = deal.stage1.rcept_no
    latest_eq_rcept = deal.stage1.rcept_no  # stage1 의 existing_qty 가 첫 값

    def _maybe_update_schedule(source_rcept, new_rd, new_pd, label):
        """더 늦은 rcept 의 non-None 일정 값으로만 갱신."""
        nonlocal latest_rd_rcept, latest_pd_rcept
        if new_rd and source_rcept > latest_rd_rcept:
            if new_rd != rec.record_date:
                print(f"  [INFO] {deal.corp_name}: 배정기준일 {rec.record_date} → {new_rd} ({label})")
            rec.record_date = new_rd
            latest_rd_rcept = source_rcept
        if new_pd and source_rcept > latest_pd_rcept:
            if new_pd != rec.payment_date:
                print(f"  [INFO] {deal.corp_name}: 납입일 {rec.payment_date} → {new_pd} ({label})")
            rec.payment_date = new_pd
            latest_pd_rcept = source_rcept

    def _maybe_update_existing_qty(source_rcept, new_eq, label):
        """더 늦은 rcept 의 non-None existing_qty 로만 갱신 (현대아산 2023.03.30 케이스).

        amend 본문 "I. 모집 또는 매출에 관한 일반사항 / 2. 공모방법" 에서 자기주식
        변동에 따른 발행주식총수 변경을 추적 → 엑셀 F (기존주식) 갱신, G (증자비율)
        는 = E/F 수식이라 자동 재계산.
        """
        nonlocal latest_eq_rcept
        if new_eq and source_rcept > latest_eq_rcept:
            if new_eq != rec.existing_qty:
                print(f"  [INFO] {deal.corp_name}: 기존주식 {rec.existing_qty} → {new_eq} ({label})")
            rec.existing_qty = new_eq
            latest_eq_rcept = source_rcept

    def _maybe_backfill_init(am_rec, label):
        """stage1 본문에 init_qty/init_price 정보 자체가 없는 특수 케이스 (WEBTOON
        2024-06 케이스) → amend 본문에서 보강. **None 일 때만** 채움.
        (한번 채워지면 더 이상 갱신 안 함 — 사용자 룰: H/I 는 최초 정보 유지.)
        """
        if rec.init_qty is None and am_rec.init_qty is not None:
            rec.init_qty = am_rec.init_qty
            print(f"  [INFO] {deal.corp_name}: init_qty {am_rec.init_qty:,} "
                  f"(amend {label} 에서 보강 — stage1 표 비어있음)")
        if rec.init_price is None and am_rec.init_price is not None:
            rec.init_price = am_rec.init_price
            print(f"  [INFO] {deal.corp_name}: init_price {am_rec.init_price:,} "
                  f"(amend {label} 에서 보강 — stage1 표 비어있음)")

    # amend 들 — 배정기준일/납입일 + 발행주식총수(자기주식) 변경 체크.
    # amend 당 2 섹션 fetch (일정 1 + 공모방법 1).
    for am in deal.amends:
        try:
            secs_am = dart_client.fetch_full_document(
                am.rcept_no,
                title_predicate=dart_client.ecm_amend_strict_predicate)
        except Exception as e:
            print(f"  [WARN] {deal.corp_name} amend {am.rcept_no}: fetch 실패 — {e}")
            continue
        am_rec = parser_ecm.parse_rights_stage1(secs_am, am.rcept_no,
                                                 deal.corp_name, deal.corp_code)
        _maybe_update_schedule(am.rcept_no, am_rec.record_date,
                                am_rec.payment_date, am.rcept_dt)
        _maybe_update_existing_qty(am.rcept_no, am_rec.existing_qty, am.rcept_dt)
        _maybe_backfill_init(am_rec, am.rcept_no)

    # final 들 (1차/2차)
    sorted_finals = sorted(deal.finals, key=lambda x: x.rcept_no)
    final1_amounts: list = []
    final2_amounts: list = []
    # 일반공모 / 제3자배정 는 단일 [발행조건확정] 만 있고 그게 곧 최종.
    # 1차/2차 단계 없음 → 단일 final 을 final2 (= 최종) 슬롯에 직접 배치.
    if rec.offering_type in ("일반공모", "제3자배정") and len(sorted_finals) >= 1:
        f = sorted_finals[0]
        rec.rcept_no_final2 = f.rcept_no
        secs_f = dart_client.fetch_full_document(
            f.rcept_no, title_predicate=dart_client.ecm_final_strict_predicate)
        result = parser_ecm.parse_rights_final(secs_f, f.rcept_no, which="stage1")
        rec.final_price = result.get("price")
        # stage1_price / stage2_price 는 None 유지 (일반공모는 단계 가액 없음)
        if result.get("updated_qty") and result["updated_qty"] != rec.new_qty:
            rec.new_qty = result["updated_qty"]
        # 일반공모: record_date 가 본문에 없으면 (배정기준일 "-") 청약기일 첫 날 사용.
        # 사용자 룰: "일반공모 유상증자의 A행은 청약기일의 첫 날".
        eff_record_date = result.get("record_date")
        if eff_record_date is None and rec.offering_type == "일반공모":
            eff_record_date = result.get("subscribe_first_date")
        _maybe_update_schedule(f.rcept_no, eff_record_date,
                                result.get("payment_date"), f"final {f.rcept_no}")
        _maybe_update_existing_qty(f.rcept_no, result.get("existing_qty"),
                                    f"final {f.rcept_no}")
        # 인수단 — 일반공모는 보통 비어있음 (증권사 안 거치는 직접 공모) 하지만 있을 수도
        final2_amounts = result.get("underwriter_amounts", [])
    elif len(sorted_finals) >= 1:
        # 1차 (주주배정 등 일반 케이스)
        f1 = sorted_finals[0]
        rec.rcept_no_final1 = f1.rcept_no
        secs_f1 = dart_client.fetch_full_document(
            f1.rcept_no, title_predicate=dart_client.ecm_final_strict_predicate)
        result1 = parser_ecm.parse_rights_final(secs_f1, f1.rcept_no, which="stage1")
        rec.stage1_price = result1.get("price")
        # 정정 후 증권수량 변동 시 갱신
        if result1.get("updated_qty") and result1["updated_qty"] != rec.new_qty:
            rec.new_qty = result1["updated_qty"]
        _maybe_update_schedule(f1.rcept_no, result1.get("record_date"),
                                result1.get("payment_date"), f"final1 {f1.rcept_no}")
        _maybe_update_existing_qty(f1.rcept_no, result1.get("existing_qty"),
                                    f"final1 {f1.rcept_no}")
        # 인수단 (1차) — 2차 없으면 이 값 사용
        final1_amounts = result1.get("underwriter_amounts", [])

    if rec.offering_type not in ("일반공모", "제3자배정") and len(sorted_finals) >= 2:
        # 2차 (마무리)
        f2 = sorted_finals[-1]
        rec.rcept_no_final2 = f2.rcept_no
        secs_f2 = dart_client.fetch_full_document(
            f2.rcept_no, title_predicate=dart_client.ecm_final_strict_predicate)
        result2 = parser_ecm.parse_rights_final(secs_f2, f2.rcept_no, which="stage2")
        rec.final_price = result2.get("price")
        rec.stage2_price = result2.get("stage2_price")
        # 정정 후 증권수량 변동 시 갱신
        if result2.get("updated_qty") and result2["updated_qty"] != rec.new_qty:
            rec.new_qty = result2["updated_qty"]
        _maybe_update_schedule(f2.rcept_no, result2.get("record_date"),
                                result2.get("payment_date"), f"final2 {f2.rcept_no}")
        _maybe_update_existing_qty(f2.rcept_no, result2.get("existing_qty"),
                                    f"final2 {f2.rcept_no}")
        final2_amounts = result2.get("underwriter_amounts", [])

    # 인수단 최종 (2차 우선, 없으면 1차) — caller 가 broker alias 집계용으로 보관
    rec._final_underwriter_amounts = final2_amounts or final1_amounts

    # stage1 fallback — final 본문에 인수단 표 없으면 stage1 본문의 인수단 표 기반으로
    # broker 비율 유지 + 정정 후 final_price (or stage1_price) 로 amount 재계산.
    # IPO 와 동일 흐름. 사용자 룰 2026-05-25.
    if not rec._final_underwriter_amounts:
        try:
            import pandas as _pd
            import io as _io
            stage1_uw = []
            for html in secs_stage1.values():
                try:
                    tables = _pd.read_html(_io.StringIO(html), flavor="lxml")
                except Exception:
                    continue
                for t in tables:
                    t = parser_ecm._promote_header(t)
                    cols = [str(c) for c in t.columns]
                    cols_joined = " ".join(cols)
                    if not (("인수인" in cols_joined or "인수(주선)인" in cols_joined)
                            and "인수수량" in cols_joined):
                        continue
                    role_col = name_col = qty_col = None
                    for ci, c in enumerate(cols):
                        cs = str(c).replace(" ", "")
                        if cs in ("인수인", "인수(주선)인") and role_col is None:
                            role_col = ci
                        elif (cs.startswith("인수인") or cs.startswith("인수(주선)인")) \
                                and name_col is None and ci != role_col:
                            name_col = ci
                        elif "인수수량" in cs:
                            qty_col = ci
                    if role_col is None or name_col is None or qty_col is None:
                        continue
                    for _, row in t.iterrows():
                        rl = str(row.iloc[role_col]).replace(" ", "")
                        nm = str(row.iloc[name_col]).replace(" ", "")
                        try:
                            qty_int = int(str(row.iloc[qty_col]).replace(",", ""))
                        except (ValueError, TypeError):
                            continue
                        if not nm or nm in ("nan", "None", "계", "합계"):
                            continue
                        stage1_uw.append({"role": rl, "name": nm, "qty": qty_int})
                    if stage1_uw:
                        break
                if stage1_uw:
                    break

            # 정정 후 최종 가액 × broker qty 비율 = amount
            # 가액 우선순위: final_price > stage2_price > stage1_price > init_price
            fp = (rec.final_price or rec.stage2_price
                  or rec.stage1_price or rec.init_price)
            # final 수량도 변경됐을 수 있음 (parse_rights_final 에서 갱신 → rec.new_qty)
            fq = rec.new_qty
            if stage1_uw and fp:
                s_total = sum(x["qty"] for x in stage1_uw) or 1
                fallback_amounts = []
                for x in stage1_uw:
                    if fq and fq != s_total:
                        new_qty = round(x["qty"] * fq / s_total)
                    else:
                        new_qty = x["qty"]
                    fallback_amounts.append({
                        "role": x["role"], "name": x["name"],
                        "qty": new_qty, "amount_won": new_qty * fp,
                    })
                rec._final_underwriter_amounts = fallback_amounts
                print(f"  [INFO] {deal.corp_name}: final 본문에 인수단 표 없음 "
                      f"→ stage1 fallback ({len(fallback_amounts)} broker, "
                      f"비율 유지 + 정정 후 가액 {fp:,}원 재계산)")
        except Exception as e:
            print(f"  [WARN] {deal.corp_name}: stage1 fallback 실패 — {e}")

    # **사용자 룰 2026-05-27**: 인수수수료 fallback (H022 자동화).
    # 위 stage1 fallback (인수수량 기준) 도 실패 = stage1 인수단 표의 "인수수량" 이
    # "-" 로 미확정인 케이스 (대한항공/한화시스템/삼성중공업 같은 대형 잔액인수 deal).
    # → 인수단 표의 "인수대가" 컬럼 (인수수수료 텍스트) 에서 broker 별 비율 추출 →
    #   모집총액 × 비율 = 각 broker amount.
    # 잔액인수 (주주배정후 실권주 일반공모 등) 케이스에 한정. role '대표' 는 lead 도
    # 함께 채워지도록 _final_underwriter_amounts 에 등록.
    if not rec._final_underwriter_amounts and \
            rec.offering_type in ("주주배정후 실권주 일반공모", "주주배정후 실권주일반공모",
                                   "주주배정", "주주배정증자"):
        try:
            fp = (rec.final_price or rec.stage2_price
                  or rec.stage1_price or rec.init_price)
            fq = rec.new_qty
            offering_total = (fp * fq) if (fp and fq) else None
            if offering_total:
                fee_brokers = parser_ecm.extract_underwriters_from_stage1_fee_pattern(
                    secs_stage1, offering_total)
                if fee_brokers:
                    rec._final_underwriter_amounts = fee_brokers
                    print(f"  [INFO] {deal.corp_name}: 인수수수료 fallback "
                          f"({len(fee_brokers)} broker, 모집총액 "
                          f"{offering_total/100000000:,.0f}억원 × 인수수수료 비율)")
        except Exception as e:
            print(f"  [WARN] {deal.corp_name}: 인수수수료 fallback 실패 — {e}")

    return rec


# ============== broker alias 집계 + 주관 실적 산식 ==============

def aggregate_broker_amounts(underwriter_amounts: list,
                              mappings: Optional[dict] = None
                              ) -> tuple[dict, set, dict]:
    """공시 본문의 인수단 행들 → (alias→억원, lead_aliases, alias→주관실적).

    role 에 "주관" 또는 "대표" 들어가면 lead.

    mappings 가 주어지면, broker_alias 매핑 실패 시 auto_register_broker_ecm 으로
    자동 등록 (DCM 동일 룰). 등록된 alias 는 mappings inplace 갱신.
    """
    eok: dict[str, float] = {}
    lead_aliases: set = set()
    for ua in underwriter_amounts:
        name = ua.get("name", "")
        role = ua.get("role", "")
        alias = parser_ecm.broker_alias(name)
        if not alias and mappings is not None and name:
            # 자동 등록 시도 — alias 생성 + config_ecm.LEAD_ECM/UW_ECM + mappings 갱신
            alias = parser_ecm.auto_register_broker_ecm(name, role, mappings)
            if alias:
                print(f"  [자동등록] '{name}' → '{alias}' (역할: {role})")
        elif alias and mappings is not None:
            # 기존 alias 라도 신규 deal 의 role 이 lead 면 LEAD_ECM 보강
            if parser_ecm.ensure_lead_alias(alias, role, mappings):
                print(f"  [lead 승격] '{alias}' (기존 UW only → LEAD 추가, "
                      f"역할: {role})")
        if not alias:
            print(f"  [WARN] alias 매칭 실패: {name!r}")
            continue
        amt = ua.get("amount_won")
        if amt is None:
            continue
        amt_eok = round(amt / 1e8, 2)
        eok[alias] = eok.get(alias, 0) + amt_eok
        # 사용자 룰 2026-05-27: '공동' = 공동주관 의미 (한화시스템 케이스).
        # '대표' / '주관' / '공동' 어느 하나라도 포함되면 lead.
        if "대표" in role or "주관" in role or "공동" in role:
            lead_aliases.add(alias)
    lead_perf = parser_ecm.compute_lead_performance(eok, lead_aliases)
    return eok, lead_aliases, lead_perf


# ============== mappings.json 로드/저장 ==============

def load_mappings() -> dict:
    """auto/mappings.json 로드 — DCM 과 공유. 없으면 빈 dict.

    또한 config_ecm.LEAD_ECM / UW_ECM 와 mappings 의 lead_managers /
    underwriters 를 merge — config_ecm.py 의 하드코딩은 baseline 이고,
    mappings 가 누적 자동등록 broker 들의 single source of truth.

    이렇게 안 하면 매 실행마다 config 가 하드코딩 25개로 reset 돼서, 자동등록된
    broker (예: 채비 케이스의 '하나') 가 시트 컬럼 인덱스 매칭에서 누락.
    """
    try:
        mappings = json.loads(config.MAPPINGS_JSON.read_text(encoding="utf-8"))
    except FileNotFoundError:
        mappings = {}
    # config_ecm 와 동기화 — mappings 의 lead_managers/underwriters 가 더 길면 추가
    for alias in mappings.get("lead_managers", []):
        if alias not in config_ecm.LEAD_ECM:
            config_ecm.LEAD_ECM.append(alias)
    for alias in mappings.get("underwriters", []):
        if alias not in config_ecm.UW_ECM:
            config_ecm.UW_ECM.append(alias)
    return mappings


def persist_auto_added(mappings: dict) -> None:
    """이번 실행에서 자동 등록된 broker 들을 mappings.json 에 영구 저장.

    DCM main._persist_auto_added 와 동일 패턴.
    """
    auto_added = mappings.pop("_auto_added_brokers", [])
    if not auto_added:
        return
    print(f"\n[자동등록] 신규 증권사 alias {len(auto_added)} 개:")
    for a in auto_added:
        print(f"  '{a['formal']}' → '{a['alias']}' (역할: {a.get('role', '')})")
    history = mappings.setdefault("_auto_added_history", [])
    history.extend(auto_added)
    config.MAPPINGS_JSON.write_text(
        json.dumps(mappings, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  → mappings.json 갱신 완료 (누적 자동등록 {len(history)} 건)")


# ============== Excel 쓰기 ==============

def write_results_to_xlsx(deals: list[DealResult], xlsx_path: Path) -> None:
    """DealResult 리스트를 시간순 정렬 후 ECM Table test.xlsx 에 쓴다.

    - IPO 시트: sort_key (상장일) 기준 오름차순, row 3 부터
    - 유상증자 시트: sort_key (배정기준일) 기준 오름차순, row 3 부터
    - 자동 등록된 신규 broker 가 있으면 두 시트 헤더 컬럼 자동 확장
    """
    import excel_writer  # DCM 의 _expand_broker_columns 재활용

    ipo_results = [r for r in deals if r and r.kind == "ipo"]
    rights_results = [r for r in deals if r and r.kind == "rights"]

    # 시간순 정렬 (None 은 맨 뒤)
    ipo_results.sort(key=lambda r: r.sort_key or date.max)
    rights_results.sort(key=lambda r: r.sort_key or date.max)

    # 파일 없으면 (= 실행 중 사용자가 지웠거나 첫 실행 외 경로) template 자동 복사.
    # 이전엔 cmd_collect 시작 시에만 체크했는데, fetch 도중 파일 삭제될 경우 save 실패 → 방어.
    if not xlsx_path.exists():
        import shutil
        shutil.copy(config_ecm.ECM_TEMPLATE, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws_ipo = wb["IPO"]
    ws_rights = wb["유상증자"]

    # 신규 broker 가 있으면 시트 헤더 컬럼 확장 (양 시트 모두)
    excel_writer._expand_broker_columns(
        ws_ipo, leads=config_ecm.LEAD_ECM, uws=config_ecm.UW_ECM)
    excel_writer._expand_broker_columns(
        ws_rights, leads=config_ecm.LEAD_ECM, uws=config_ecm.UW_ECM)

    # 기존 데이터 영역 초기화 (row 3 부터 끝까지)
    for ws in (ws_ipo, ws_rights):
        for r in range(3, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None

    # IPO 쓰기
    for idx, dr in enumerate(ipo_results):
        _write_ipo_row(ws_ipo, 3 + idx, dr)
    # 유상증자 쓰기 — multi-issue 시 row 카운터 누적 (한 deal 이 2+ row 차지 가능)
    r = 3
    for dr in rights_results:
        last_r = _write_rights_row(ws_rights, r, dr)
        r = last_r + 1

    wb.save(xlsx_path)

    # **사이드카 meta.json 저장** — row 별 rcept_no 매핑.
    # validator_ecm 의 refetch 핸들러가 이 매핑으로 즉시 rcept_no 받아 효율 처리.
    meta_path = xlsx_path.with_suffix(".meta.json")
    meta = {
        "ipo_records": [
            {
                "row_idx": 3 + i,
                "corp_name": dr.ipo_record.issuer if dr.ipo_record else "",
                "corp_code": dr.ipo_record.corp_code if dr.ipo_record else "",
                "rcept_no_stage1": dr.ipo_record.rcept_no_stage1 if dr.ipo_record else "",
                "rcept_no_final":  dr.ipo_record.rcept_no_final  if dr.ipo_record else "",
            }
            for i, dr in enumerate(ipo_results)
        ],
        "rights_records": [
            {
                "row_idx": 3 + i,
                "corp_name": dr.rights_record.issuer if dr.rights_record else "",
                "corp_code": dr.rights_record.corp_code if dr.rights_record else "",
                "rcept_no_stage1": dr.rights_record.rcept_no_stage1 if dr.rights_record else "",
                "rcept_no_final1": dr.rights_record.rcept_no_final1 if dr.rights_record else "",
                "rcept_no_final2": dr.rights_record.rcept_no_final2 if dr.rights_record else "",
            }
            for i, dr in enumerate(rights_results)
        ],
    }
    meta_path.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → meta.json: {meta_path.name} ({len(meta['ipo_records'])} IPO + "
          f"{len(meta['rights_records'])} 유상증자 row 매핑)")


def _copy_row_format(ws, src_row: int, dst_row: int) -> None:
    """src_row 의 모든 셀 포맷 (number_format, font, alignment, fill, border) 을
    dst_row 로 복사. 향후 row 가 중간에 insert 되는 경우, 또는 데이터 추가 시
    포맷 일관성 보장용. 셀 값은 건드리지 않음.

    호출 시점:
      - insert_rows 직후
      - 데이터 영역 외 행으로 데이터를 새로 쓸 때 (예: 템플릿 max_row 초과)
    """
    from copy import copy as _copy
    if src_row == dst_row:
        return
    for c in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        dst.number_format = src.number_format
        # 셀 스타일 — 깊은 복사 (StyleArray 가 immutable 이라 직접 할당 가능)
        if src.has_style:
            dst.font = _copy(src.font)
            dst.alignment = _copy(src.alignment)
            dst.fill = _copy(src.fill)
            dst.border = _copy(src.border)
            dst.protection = _copy(src.protection)


def _ensure_row_format(ws, row: int, ref_row: int = 3) -> None:
    """행 row 의 포맷이 ref_row (보통 row 3) 와 다르면 ref_row 의 포맷으로 맞춤.
    데이터를 쓰기 전에 호출해서, 어떤 경로로 행이 생성됐든 일관성 보장.

    체크 항목:
      - number_format (G='General' → 표준)
      - alignment.horizontal (dedup/sort 후 A~C 좌측정렬로 풀리는 케이스 방지)
    """
    if row == ref_row:
        return
    from copy import copy as _copy
    for c in range(1, ws.max_column + 1):
        ref_cell = ws.cell(row=ref_row, column=c)
        cur_cell = ws.cell(row=row, column=c)
        if cur_cell.number_format != ref_cell.number_format:
            cur_cell.number_format = ref_cell.number_format
        # alignment.horizontal 일치 보장 (vertical/wrap_text 는 영향 적어 생략)
        ref_a = ref_cell.alignment
        cur_a = cur_cell.alignment
        if ref_a is not None and (cur_a is None or cur_a.horizontal != ref_a.horizontal):
            cur_cell.alignment = _copy(ref_a)


def _write_ipo_row(ws, row: int, dr: DealResult) -> None:
    rec = dr.ipo_record
    if rec is None:
        return
    # 데이터 행 포맷을 row 3 (template 첫 행) 과 동기화 — row insert 또는 외부
    # 편집으로 포맷이 묶인 경우 (예: 대한광통신 G='General') 자동 정정.
    _ensure_row_format(ws, row, ref_row=3)
    # 상장일 우선순위: report 확정 상장일 > KIND 예정일 > "미정"
    # (process_deal 에서 이미 우선순위 적용됨 → _listing_date 가 None 이면 둘 다 실패)
    listing = getattr(rec, "_listing_date", None) or dr.listing_date_planned
    ws.cell(row=row, column=1).value = listing if listing else "미정"
    ws.cell(row=row, column=2).value  = rec.issuer
    ws.cell(row=row, column=3).value  = rec.market
    ws.cell(row=row, column=4).value  = rec.init_qty
    ws.cell(row=row, column=5).value  = rec.init_price
    # F (최초 총액), I (최종 총액) — 억원 단위 정수 (ROUND)
    ws.cell(row=row, column=6).value  = f"=ROUND(D{row}*E{row}/100000000,0)"
    # **사용자 룰 2026-05-29**: 최종 컬럼(G 최종수량 / H 최종가액 / I 최종총액)은
    # [발행조건확정] 도달(rcept_no_final 존재) 시에만 기입. stage1 단계에선 최초희망수량을
    # 최종수량으로 미리 복사하지 않고 비워둠 (확정 전 데이터를 확정처럼 보이지 않게).
    # 확정 단계의 G=final_qty 는 parse_ipo_final 이 '정정 후' 모집표 증권수량에서 실제 추출.
    if rec.rcept_no_final:
        ws.cell(row=row, column=7).value  = rec.final_qty or rec.init_qty
        ws.cell(row=row, column=8).value  = rec.final_price
        ws.cell(row=row, column=9).value  = f"=ROUND(G{row}*H{row}/100000000,0)"
    else:
        ws.cell(row=row, column=7).value  = None
        ws.cell(row=row, column=8).value  = None
        ws.cell(row=row, column=9).value  = None
    ws.cell(row=row, column=10).value = rec.new_share_ratio
    ws.cell(row=row, column=11).value = f"=1-J{row}"
    ws.cell(row=row, column=12).value = rec.inst_initial
    ws.cell(row=row, column=13).value = rec.inst_subscribed
    if rec.inst_initial and rec.inst_subscribed:
        ws.cell(row=row, column=14).value = f"=M{row}/L{row}"
    ws.cell(row=row, column=15).value = rec.inst_final
    ws.cell(row=row, column=16).value = rec.general_initial
    ws.cell(row=row, column=17).value = rec.general_subscribed
    if rec.general_initial and rec.general_subscribed:
        ws.cell(row=row, column=18).value = f"=Q{row}/P{row}"
    ws.cell(row=row, column=19).value = rec.general_final
    ws.cell(row=row, column=20).value = rec.esop_initial
    ws.cell(row=row, column=21).value = rec.esop_final
    # V (우리사주_청약률) — esop_initial > 0 이면 V 수식 채움.
    # esop_final = 0 (청약 zero) 도 정상 — 자동 청약률 0% 계산.
    # 사용자 룰 2026-05-27 (정정): 핀텔 케이스 (init=100k, final=0).
    if rec.esop_initial and rec.esop_final is not None:
        ws.cell(row=row, column=22).value = f"=U{row}/T{row}"

    # 주관/인수 영역 (LEAD_ECM/UW_ECM 통합 리스트) — 억원 단위, 사용자 요구로 정수 표기
    lead_col_start = config_ecm.COL_IPO["주관_시작"]
    uw_col_start   = lead_col_start + len(config_ecm.LEAD_ECM)
    for i, broker in enumerate(config_ecm.LEAD_ECM):
        v = dr.lead_perf.get(broker)
        if v is not None and v > 0:
            ws.cell(row=row, column=lead_col_start + i).value = round(v)
    for i, broker in enumerate(config_ecm.UW_ECM):
        v = dr.underwriter_amounts_eok.get(broker)
        if v is not None and v > 0:
            ws.cell(row=row, column=uw_col_start + i).value = round(v)


def _write_rights_row(ws, row: int, dr: DealResult) -> int:
    """유증 row 작성. multi-issue 시 메인 row + extra row N개 추가 작성.
    Returns: 작성한 **마지막** row index (caller 가 다음 row 결정용).
    """
    rec = dr.rights_record
    if rec is None:
        return row
    # 데이터 행 포맷을 row 3 과 동기화 — 어떤 경로로 row 가 들어왔든 일관성 보장.
    _ensure_row_format(ws, row, ref_row=3)
    ws.cell(row=row, column=1).value  = rec.record_date
    ws.cell(row=row, column=2).value  = rec.issuer
    ws.cell(row=row, column=3).value  = rec.offering_type
    ws.cell(row=row, column=4).value  = rec.payment_date
    ws.cell(row=row, column=5).value  = rec.new_qty
    ws.cell(row=row, column=6).value  = rec.existing_qty
    if rec.new_qty and rec.existing_qty:
        g_cell = ws.cell(row=row, column=7)
        g_cell.value = f"=E{row}/F{row}"
        # row insert 등으로 number_format 이 'General' 로 묶이는 경우 방지
        g_cell.number_format = "0.00%"
    # **사용자 룰 (2026-05-25)**: H (최초_수량) / I (최초_가액) 는 정정 단계에서
    # 변경 안 됨. parser_ecm.parse_rights_stage1 의 init_qty 에 stage1 시점의 값을
    # 별도 저장 — 이후 amend/final 단계에서 rec.new_qty 만 갱신, init_qty 는 불변.
    ws.cell(row=row, column=8).value  = rec.init_qty
    ws.cell(row=row, column=9).value  = rec.init_price
    # J (최초 총액) / L (1차 총액) / N (2차 총액) / P (최종 총액) — 억원 단위 정수 (ROUND)
    if rec.init_qty and rec.init_price:
        ws.cell(row=row, column=10).value = f"=ROUND(H{row}*I{row}/100000000,0)"
    if rec.stage1_price:
        ws.cell(row=row, column=11).value = rec.stage1_price
        ws.cell(row=row, column=12).value = f"=ROUND(E{row}*K{row}/100000000,0)"
    if rec.stage2_price:
        ws.cell(row=row, column=13).value = rec.stage2_price
        ws.cell(row=row, column=14).value = f"=ROUND(E{row}*M{row}/100000000,0)"
    if rec.final_price:
        ws.cell(row=row, column=15).value = rec.final_price
        ws.cell(row=row, column=16).value = f"=ROUND(E{row}*O{row}/100000000,0)"

    # 주관/인수 영역 (LEAD_ECM/UW_ECM 통합 리스트) — 억원 단위, 정수 표기
    lead_col_start = config_ecm.COL_RIGHTS["주관_시작"]
    uw_col_start   = lead_col_start + len(config_ecm.LEAD_ECM)
    for i, broker in enumerate(config_ecm.LEAD_ECM):
        v = dr.lead_perf.get(broker)
        if v is not None and v > 0:
            ws.cell(row=row, column=lead_col_start + i).value = round(v)
    for i, broker in enumerate(config_ecm.UW_ECM):
        v = dr.underwriter_amounts_eok.get(broker)
        if v is not None and v > 0:
            ws.cell(row=row, column=uw_col_start + i).value = round(v)

    # ============ multi-issue 추가 row (솔루스첨단소재 2022-04 케이스) ============
    # rec._extra_issues 가 있으면 메인 row 외 N row 추가. 공통 메타 (기준일/회사명/구분
    # /납입일) + 자기 issue 의 모집수량/최초가액. F (기존주식), K/M/O (1차/2차/최종
    # 가액), broker — H022/H009/H010/H011 룰로 사후 보정.
    extras = getattr(rec, "_extra_issues", None) or []
    last_row = row
    for extra in extras:
        last_row += 1
        _ensure_row_format(ws, last_row, ref_row=3)
        # 공통 메타 (메인 row 와 동일)
        ws.cell(row=last_row, column=1).value = rec.record_date
        ws.cell(row=last_row, column=2).value = rec.issuer  # 같은 회사명 (사용자 룰)
        ws.cell(row=last_row, column=3).value = rec.offering_type
        ws.cell(row=last_row, column=4).value = rec.payment_date
        # E 모집수량
        e_val = extra.get("qty")
        ws.cell(row=last_row, column=5).value = e_val
        # F 기존주식 / G 증자비율 — 보강은 H022 룰
        # H 최초수량 / I 최초가액 / J 최초총액
        ws.cell(row=last_row, column=8).value = e_val
        ws.cell(row=last_row, column=9).value = extra.get("init_price")
        if e_val and extra.get("init_price"):
            ws.cell(row=last_row, column=10).value = f"=ROUND(H{last_row}*I{last_row}/100000000,0)"
        # K/M/O (1차/2차/최종 가액) — Step 3 향후 자동화. 현재는 H009/H010/H011 룰로 보정.
        # broker — 메인 row 의 broker 와 다른 분배 필요 (자기 issue 의 모집총액 기준).
        # 현재 비움. H022 룰로 사후 보정.

    return last_row


# ============== 메인 명령 ==============

def cmd_collect(start: date, end: date, corp_code: str = "",
                xlsx: Optional[Path] = None,
                run_validator: bool = True) -> None:
    """기간 내 ECM 공시 수집 → 파싱 → 저장 → validator 자동 검증·수정.

    DCM cmd_collect/cmd_update 와 동일 패턴 — 수집 후 validator 자동 실행으로
    Layer 1 룰 검증 + auto-fix patch + Pass 2.4 dedup + Pass 2.5 sort 까지 한 번에.
    """
    print(f"=== ECM 공시 수집 {start} ~ {end} ===")

    xlsx = xlsx or config_ecm.ECM_XLSX_TEST
    if not xlsx.exists():
        # 템플릿 복사
        import shutil
        shutil.copy(config_ecm.ECM_TEMPLATE, xlsx)
        print(f"  템플릿 복사: {config_ecm.ECM_TEMPLATE.name} → {xlsx.name}")

    # mappings.json (DCM 과 공유) — 자동 broker 등록에 사용
    mappings = load_mappings()

    # 1. **primary fetch**: 사용자 기간 안에 등장한 ECM 공시 목록.
    print(f"  [1] primary fetch (사용자 기간): {start} ~ {end}")
    primary = dart_client.list_ecm_filings(start, end, corp_code=corp_code)
    print(f"      ECM 공시 {len(primary)} 건")

    # 2. 분류 통계
    by_kind = {"stage1": 0, "stage1_backfill": 0, "amend": 0,
               "final": 0, "report": 0, "ignore": 0}
    for f in primary:
        by_kind[parser_ecm.classify_filing(f.report_nm)] += 1
    for k, v in by_kind.items():
        print(f"    {k}: {v}")

    # 3. 딜 그룹화 — stage1 없는 corp (amend/final/report 만 있는 corp) 도 임시 deal 생성
    deals = group_into_deals(primary)
    print(f"\n  [2] 딜 그룹: {len(deals)} 개 (사용자 기간 활성)")

    # 4. **lazy backfill** — 진짜 stage1 없거나 final 일부 누락된 deal 을 dropdown 으로 백필
    #    "같은 한 묶음 공시" (main.do dropdown) 에서 진짜 stage1 + 누락 final 모두 받기.
    #    유상증자는 final1/final2 두 단계 있는데 final1 이 사용자 기간 밖이면 누락 가능 →
    #    이 경우도 dropdown 으로 받아 process_rights 가 final1/final2 모두 정상 처리.
    needs_bf = []
    for d in deals:
        if d.stage1 is None:
            needs_bf.append(d)
        elif parser_ecm.classify_filing(d.stage1.report_nm) != "stage1":
            # stage1_backfill (= [정정제출요구]/[첨부추가]) — 진짜 stage1 로 교체 시도
            needs_bf.append(d)
        elif len(d.finals) == 1:
            # 유상증자의 final1 누락 가능성 — dropdown 으로 final 전부 확인
            needs_bf.append(d)
    print(f"  [3] dropdown 백필 필요: {len(needs_bf)} deal")

    bf_ok = 0
    bf_fail = []
    for d in needs_bf:
        # 어떤 공시 rcept_no 든 한 건으로 dropdown 받기
        sample = d.stage1
        if sample is None:
            for arr in (d.amends, d.finals, d.reports):
                if arr:
                    sample = arr[0]
                    break
        if sample is None:
            bf_fail.append(d)
            continue
        try:
            dropdown = dart_client.fetch_deal_filings(sample.rcept_no)
        except Exception as e:
            print(f"      [WARN] {d.corp_name} dropdown 실패: {e}")
            bf_fail.append(d)
            continue
        # 진짜 stage1 (접두어 없음) 찾기 — dropdown 의 가장 이른 stage1
        true_stage1 = None
        for f in dropdown:
            if parser_ecm.classify_filing(f.report_nm) == "stage1":
                true_stage1 = f
                break
        # 누락된 final 도 함께 추가 (유상증자 final1 누락 케이스 대응)
        existing_final_rcepts = {ff.rcept_no for ff in d.finals}
        new_finals = []
        for f in dropdown:
            if f.rcept_no in existing_final_rcepts:
                continue
            # dropdown 의 [발행조건확정] 표시는 단순 텍스트라 classify 가 final 로 잡힘
            if parser_ecm.classify_filing(f.report_nm) == "final":
                f.corp_name = d.corp_name
                f.corp_code = d.corp_code
                new_finals.append(f)
        if new_finals:
            d.finals.extend(new_finals)
            d.finals.sort(key=lambda x: x.rcept_no)
            print(f"      [+final] {d.corp_name}: dropdown 에서 final {len(new_finals)}개 추가")

        if true_stage1 is None:
            # final 만 추가됐어도 stage1 없으면 skip
            if d.stage1 is None:
                bf_fail.append(d)
            continue
        # dropdown 에는 corp_name/corp_code 가 없으므로 deal 에서 채움
        true_stage1.corp_name = d.corp_name
        true_stage1.corp_code = d.corp_code
        # 기존 stage1 (backfill 자리) 이 있으면 amends 로 이동
        if d.stage1 is not None and d.stage1.rcept_no != true_stage1.rcept_no:
            d.amends.insert(0, d.stage1)
        d.stage1 = true_stage1
        bf_ok += 1
    print(f"      백필 성공: {bf_ok} / 실패: {len(bf_fail)}")
    for d in bf_fail:
        print(f"        [SKIP] {d.corp_name}: 진짜 stage1 못 찾음")

    # 5. stage1 여전히 None 인 deal 제외 (처리 불가)
    deals = [d for d in deals if d.stage1 is not None]
    print(f"\n  최종 처리 대상 딜: {len(deals)} 개")

    # 4. 각 딜 처리
    results: list[DealResult] = []
    for i, deal in enumerate(deals, 1):
        if deal.stage1 is None:
            continue
        # 사용자 룰 2026-05-26: 철회된 deal 은 수집 대상 아님
        if deal.is_withdrawn:
            print(f"\n  [{i}/{len(deals)}] {deal.corp_name} (stage1={deal.stage1.rcept_no}) "
                  f"— [철회 skip]")
            continue
        print(f"\n  [{i}/{len(deals)}] {deal.corp_name} (stage1={deal.stage1.rcept_no})"
              f"  amends={len(deal.amends)} finals={len(deal.finals)} reports={len(deal.reports)}")
        try:
            res = process_deal(deal)
        except Exception as e:
            print(f"    [ERROR] process_deal 실패: {e}")
            import traceback; traceback.print_exc()
            continue
        if res is None:
            continue
        # broker alias 집계 (mappings 전달 → 자동 등록 흐름 활성화)
        if res.kind == "ipo" and res.ipo_record is not None:
            # 사용자 룰 2026-05-25: IPO broker amount 는 **첫 번째 [발행조건확정] 본문**에서.
            # 보고서 (_report_underwriter_amounts) 는 청약결과 추출용으로만 사용.
            # _final_underwriter_rows 는 이미 amount_won 포함 (parse_ipo_final 강화 + stage1 fallback).
            final_rows = getattr(res.ipo_record, "_final_underwriter_rows", [])
            eok, leads, perf = aggregate_broker_amounts(final_rows, mappings=mappings)
            res.underwriter_amounts_eok = eok
            res.lead_aliases = leads
            res.lead_perf = perf
            res.listing_date_planned = getattr(res.ipo_record, "_listing_date", None)
            res.sort_key = res.listing_date_planned
        elif res.kind == "rights" and res.rights_record is not None:
            uw_rows = getattr(res.rights_record, "_final_underwriter_amounts", [])
            eok, leads, perf = aggregate_broker_amounts(uw_rows, mappings=mappings)
            res.underwriter_amounts_eok = eok
            res.lead_aliases = leads
            res.lead_perf = perf
            res.sort_key = res.rights_record.record_date

        results.append(res)
        print(f"    [OK] {res.kind} — sort_key={res.sort_key}")

    # 5. 엑셀 쓰기 (시간순 정렬 + 새 broker 컬럼 자동 추가)
    print(f"\n  >>> {xlsx.name} 에 {len(results)} 건 쓰기...")
    write_results_to_xlsx(results, xlsx)
    print(f"  [완료]")

    # 6. 자동 등록된 broker 영구 저장
    persist_auto_added(mappings)

    # 7. validator 자동 실행 (사용자 룰 2026-05-25: DCM 동일 — 수집·검증·수정 한 묶음)
    if run_validator:
        print(f"\n=== validator 자동 실행 ===")
        import subprocess
        subprocess.run(
            [sys.executable, "-X", "utf8", str(config.ROOT / "validator_ecm.py"),
             "--xlsx", str(xlsx)],
            check=False, encoding="utf-8")


def _finalize_overdue_rights_deals(xlsx_path: Path,
                                   merged_meta: dict,
                                   today: Optional[date] = None,
                                   wide_scan: bool = False) -> int:
    """납입일이 지났지만 [발행조건확정] 없이 마무리된 유상증자 딜 자동 마감.

    현대아산 2023-04 케이스: [발행조건확정] 공시 없이 납입일에 곧바로
    "증권발행실적보고서" 가 올라옴. 묶음 공시의 가장 늦은 정정 (또는 stage1)
    가액으로 O (최종_가액) 채우고 딜 마감.

    조건:
      - 유상증자 row 중 O (최종_가액) 비어있음
      - D (납입일) <= today (납입일 도래)
      - (wide_scan=False, 기본) today <= 납입일+7일 — 매일 cmd_update 모니터링 윈도우
      - (wide_scan=True) 윈도우 검사 skip — 일회성 backfill 용 (오래된 케이스도 포함)

    동작:
      1. fetch_deal_filings 로 묶음 공시 조회
      2. "증권발행실적보고서" 가 (납입일-1 ~ 납입일+7) 범위에 있는지 확인
      3. 있으면 → 묶음 내 [발행조건확정]·실적보고서 제외한 가장 늦은 공시
         (보통 마지막 [정정] 또는 stage1) 본문 fetch
      4. parse_offering_summary → init_price 추출 → O 셀 patch
      5. meta record 의 rcept_no_report 슬롯에 보고서 rcept_no 기록

    Returns: 마감 처리된 row 개수.
    """
    from datetime import timedelta
    if today is None:
        today = date.today()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["유상증자"]

    # row_idx → meta record 매핑
    meta_by_row = {rec.get("row_idx"): rec
                   for rec in merged_meta.get("rights_records", [])}

    # 1) 후보 row 추출
    candidates = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value
        pay = ws.cell(r, 4).value   # D = 납입일
        final = ws.cell(r, 15).value  # O = 최종_가액
        if not name or pay is None or final not in (None, ""):
            continue
        # 날짜 타입 정규화
        pay_date = pay.date() if hasattr(pay, "date") else pay
        if not isinstance(pay_date, date):
            continue
        # 윈도우 검사 — wide_scan 시 skip
        if not wide_scan:
            if not (pay_date <= today <= pay_date + timedelta(days=7)):
                continue
        else:
            if not (pay_date <= today):
                continue
        rec = meta_by_row.get(r)
        if not rec or not rec.get("rcept_no_stage1"):
            continue
        # [발행조건확정] 공시가 이미 등록된 케이스는 _finalize 대상 아님 — parser_ecm.
        # parse_rights_final 가 정정 후 가액 추출 못한 별개 버그이므로, _finalize 가
        # 잘못 잡으면 stage1/amend 의 정정 전 가액을 채워넣어 오류 (OCI홀딩스 2023-08 케이스).
        if rec.get("rcept_no_final1") or rec.get("rcept_no_final2"):
            continue
        candidates.append((r, name, pay_date, rec))

    if not candidates:
        return 0

    print(f"\n=== 납입일 도래 + [발행조건확정] 부재 유증 딜 follow-up ===")
    _mode_label = "전체 과거" if wide_scan else "납입일 ~ +7일 윈도우"
    print(f"  후보 {len(candidates)}건 ({_mode_label})")

    finalized = 0
    for r, name, pay_date, rec in candidates:
        stage1_rcept = rec["rcept_no_stage1"]
        corp_code = rec.get("corp_code", "")

        # 1) 증권발행실적보고서 — corp_code 로 별도 검색 (트리거 신호)
        # fetch_deal_filings 의 묶음 dropdown 에는 실적보고서가 포함되지 않으므로
        # OpenDART list_ecm_filings 로 corp_code + 납입일 ±7일 범위 query 필요.
        if not corp_code:
            print(f"  [SKIP] {name} (row {r}) corp_code 없음 → 검색 불가")
            continue
        try:
            report_filings = dart_client.list_ecm_filings(
                start=pay_date - timedelta(days=1),
                end=pay_date + timedelta(days=7),
                corp_code=corp_code)
        except Exception as e:
            print(f"  [WARN] {name} (row {r}) 실적보고서 검색 실패: {e}")
            continue

        report = None
        for f in report_filings:
            if "증권발행실적보고서" in (f.report_nm or "").replace(" ", ""):
                report = f
                break

        if not report:
            print(f"  [PENDING] {name} (row {r}) 실적보고서 아직 없음 — 윈도우 내 재시도")
            continue

        # 2) 묶음 dropdown → 가장 늦은 정정 (또는 stage1) 찾기 (가액 추출용)
        try:
            filings = dart_client.fetch_deal_filings(stage1_rcept)
        except Exception as e:
            print(f"  [WARN] {name} (row {r}) 묶음 공시 조회 실패: {e}")
            continue

        last_amend = None
        for f in filings:
            nm = f.report_nm.replace(" ", "")
            if "발행조건확정" in nm or "증권발행실적보고서" in nm:
                continue
            if "증권신고서(지분증권)" in nm:
                last_amend = f  # rcept_no asc 정렬이므로 마지막 매칭 유지

        target_rcept = last_amend.rcept_no if last_amend else stage1_rcept

        # 본문 fetch — stage1 strict (모집표 + 공모방법 2 섹션)
        try:
            secs = dart_client.fetch_full_document(
                target_rcept,
                title_predicate=dart_client.ecm_stage1_strict_predicate)
        except Exception as e:
            print(f"  [WARN] {name} (row {r}) {target_rcept} fetch 실패: {e}")
            continue

        summ = parser_ecm.parse_offering_summary(secs)
        final_price = summ.get("init_price")
        if not final_price:
            print(f"  [WARN] {name} (row {r}) {target_rcept} init_price 추출 실패")
            continue

        # Swap sanity check — amend 본문에서 total 미추출 시 parser swap_fix 가 작동 못함.
        # 엑셀 E (수량) 기반으로 1주당 가액 비현실적이면 swap 가설 적용.
        qty_e = ws.cell(r, 5).value
        if (isinstance(final_price, int) and final_price > 1_000_000
                and isinstance(qty_e, (int, float)) and qty_e > 0):
            cand = final_price / qty_e
            if abs(cand - round(cand)) < 0.5:
                cand_int = int(round(cand))
                if 1 <= cand_int <= 1_000_000:
                    print(f"    [SWAP_FIX] {name}: {final_price:,} → {cand_int:,}원 "
                          f"(수량 {int(qty_e):,} 로 나눔)")
                    final_price = cand_int

        # O 셀 patch
        ws.cell(r, 15).value = final_price
        # P (최종_총액) — 수식이 비어 있으면 추가 (cmd_collect 가 O 가 비었을 때 안 넣은 케이스)
        p_cell = ws.cell(r, 16)
        if p_cell.value in (None, ""):
            p_cell.value = f"=ROUND(E{r}*O{r}/100000000,0)"
            # 옆 row 스타일 복제 — 가장 가까운 비어있지 않은 P 셀
            for src_r in (r - 1, r + 1, r - 2, r + 2):
                if 3 <= src_r <= ws.max_row:
                    src = ws.cell(src_r, 16)
                    if src.value:
                        from copy import copy as _copy
                        p_cell.font = _copy(src.font)
                        p_cell.alignment = _copy(src.alignment)
                        p_cell.number_format = src.number_format
                        break
        rec["rcept_no_report"] = report.rcept_no
        print(f"  [FIX] {name} (row {r}): O (최종_가액) = {final_price:,}원 "
              f"(source rcept={target_rcept}, report={report.rcept_no})")
        finalized += 1

    if finalized:
        wb.save(xlsx_path)
    print(f"  [완료] {finalized}건 마감")
    return finalized


def _finalize_overdue_ipo_reports(xlsx_path: Path,
                                  merged_meta: dict,
                                  today: Optional[date] = None) -> int:
    """**사용자 룰 2026-05-29 (IPO 전용)**:
    [발행조건확정] 까지 진행됐으나 '증권발행실적보고서' 가 아직 수집 안 된 IPO 딜에
    대해, cmd_update 시 자동으로 보고서 게시 여부 재확인 + 데이터 마무리.

    - 룰 "신규 공시와 무관한 이전 데이터 손대지 말라" 의 명시적 예외 — 사용자 허용.
      (유증은 _finalize_overdue_rights_deals 영구 비활성 유지)
    - 옵션 A: 상장일이 보고서값과 다르면 그냥 덮어쓰기 + INFO log
    - 옵션 A: cmd_update 자동 따라붙기 (별도 wide_scan 명령 없음)

    트리거 조건 (AND):
      - IPO meta record 中 rcept_no_report 비어있음
      - rcept_no_final 존재 (= [발행조건확정] 도달)
      - **엑셀 데이터 상 미마감** (아래 셋 중 하나라도 해당):
          a) A (상장일) 가 None or "미정"
          b) L (기관 최초배정) 비어있음 (청약/배정 결과 미확정)
          c) 주관/인수 영역 전체 비어있음 (H025 케이스)
        → 사용자 룰 "신규와 무관한 이전 데이터 손대지 말라" 의 안전장치.
          위 셋 모두 채워진 row 는 meta 의 rcept_no_report 슬롯이 비어있어도
          이미 마감된 deal 로 간주하고 손대지 않음.
      - 엑셀 A (상장일) <= today + 7  (예정일 도래 또는 임박)
        — A가 None / "미정" 인 경우는 rcept_no_final 공시일 + 60일 이내로 fallback

    동작:
      1. corp_code 로 (예정일 - 7) ~ (today + 7) 범위에서 증권발행실적보고서 검색
      2. 발견 시 본문 fetch + parse_ipo_report
      3. 셀 patch:
          - A (col 1) : listing_date_confirmed (변경 시 INFO log)
          - L/M/O    : inst_initial / inst_subscribed / inst_final
          - P/Q/S    : general_initial / general_subscribed / general_final
          - T/U      : esop_initial / esop_final
          - N/R/V    : 경쟁률 / 청약률 수식 (자동)
          - 주관/인수 (W~) : 기존 셀이 빈 경우 report 의 underwriter_amounts fallback
       4. meta record rcept_no_report 슬롯 기록

    Returns: 마감 처리된 row 개수.
    """
    from datetime import timedelta
    if today is None:
        today = date.today()

    wb = openpyxl.load_workbook(xlsx_path)
    if "IPO" not in wb.sheetnames:
        return 0
    ws = wb["IPO"]

    meta_by_row = {rec.get("row_idx"): rec
                   for rec in merged_meta.get("ipo_records", [])}

    # 1) 후보 추출
    lead_col_start = config_ecm.COL_IPO["주관_시작"]
    uw_col_end = lead_col_start + len(config_ecm.LEAD_ECM) + len(config_ecm.UW_ECM)

    candidates = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        rec = meta_by_row.get(r)
        if not rec:
            continue
        # **사용자 룰 2026-05-29**: 수동 확정(verified) row 는 완전 보호 — follow-up skip
        if rec.get("verified"):
            continue
        # 이미 보고서 fetch 완료
        if rec.get("rcept_no_report"):
            continue
        # [발행조건확정] 미도달 — 보고서 검색 의미 없음
        if not rec.get("rcept_no_final"):
            continue
        # corp_code 없으면 검색 불가
        if not rec.get("corp_code"):
            continue

        # **엑셀 데이터 상 미마감 여부** 점검 — 사용자 룰 "이전 데이터 손대지 말라" 안전장치
        a_val = ws.cell(r, 1).value
        l_val = ws.cell(r, 12).value
        a_unfinished = (a_val is None or a_val == "" or a_val == "미정")
        l_empty = (l_val is None or l_val == "")
        brokers_empty = all(
            ws.cell(r, ci).value in (None, "")
            for ci in range(lead_col_start, uw_col_end)
        )
        if not (a_unfinished or l_empty or brokers_empty):
            # 엑셀 상 이미 마감된 deal — meta 의 rcept_no_report 슬롯만 비어있는
            # historical artifact. 손대지 않음.
            continue

        # 예정 상장일 vs today
        listing_planned = None
        if hasattr(a_val, "date"):
            listing_planned = a_val.date()
        elif isinstance(a_val, date):
            listing_planned = a_val
        # 윈도우 결정
        if listing_planned:
            if listing_planned > today + timedelta(days=7):
                continue  # 너무 미래 → skip
            search_start = listing_planned - timedelta(days=7)
            search_end = min(today + timedelta(days=7), listing_planned + timedelta(days=90))
        else:
            # 예정일 미정 — final 공시일 + 60일 윈도우
            f_rcept = rec.get("rcept_no_final", "")
            if len(f_rcept) < 8:
                continue
            try:
                f_date = date(int(f_rcept[:4]), int(f_rcept[4:6]), int(f_rcept[6:8]))
            except ValueError:
                continue
            if f_date + timedelta(days=60) < today:
                continue  # 60일 지났는데 안 올라옴 → 이 패스에서 skip (다음 패스에서 재시도)
            search_start = f_date
            search_end = min(today + timedelta(days=7), f_date + timedelta(days=90))
        candidates.append((r, name, listing_planned, search_start, search_end, rec))

    if not candidates:
        return 0

    print(f"\n=== IPO 미마감 딜 증권발행실적보고서 follow-up ===")
    print(f"  후보 {len(candidates)}건")

    finalized = 0
    for r, name, listing_planned, search_start, search_end, rec in candidates:
        corp_code = rec["corp_code"]
        # 1) 보고서 검색
        try:
            filings = dart_client.list_ecm_filings(
                start=search_start, end=search_end, corp_code=corp_code)
        except Exception as e:
            print(f"  [WARN] {name} (row {r}) 보고서 검색 실패: {e}")
            continue

        report_f = None
        for f in filings:
            if "증권발행실적보고서" in (f.report_nm or "").replace(" ", ""):
                report_f = f
                break

        if not report_f:
            print(f"  [PENDING] {name} (row {r}) 실적보고서 아직 없음")
            continue

        # 2) 본문 fetch + parse
        try:
            secs = dart_client.fetch_full_document(
                report_f.rcept_no,
                title_predicate=dart_client.ecm_report_strict_predicate)
        except Exception as e:
            print(f"  [WARN] {name} (row {r}) report {report_f.rcept_no} fetch 실패: {e}")
            continue

        report = parser_ecm.parse_ipo_report(secs, report_f.rcept_no)

        # 3) 상장일 — 변경 시 INFO log + 덮어쓰기
        listing_confirmed = report.get("listing_date_confirmed")
        if listing_confirmed and rec.get("rcept_no_stage1"):
            listing_confirmed = parser_ecm._maybe_fix_listing_date_typo(
                listing_confirmed, rec["rcept_no_stage1"])
        if listing_confirmed:
            if listing_planned and listing_confirmed != listing_planned:
                print(f"  [INFO] {name}: 상장일 {listing_planned} → {listing_confirmed} "
                      f"(report {report_f.rcept_no})")
            ws.cell(r, 1).value = listing_confirmed

        # 4) 청약/배정 셀 patch — 보고서값 있을 때만
        def _set_if(col, key):
            v = report.get(key)
            if v is not None:
                ws.cell(r, col).value = v

        _set_if(12, "inst_initial")     # L
        _set_if(13, "inst_subscribed")  # M
        _set_if(15, "inst_final")       # O
        _set_if(16, "general_initial")  # P
        _set_if(17, "general_subscribed")  # Q
        _set_if(19, "general_final")    # S
        _set_if(20, "esop_initial")     # T
        _set_if(21, "esop_final")       # U

        # 경쟁률/청약률 수식 — 분자/분모 다 채워졌으면 자동 채움
        if report.get("inst_initial") and report.get("inst_subscribed"):
            ws.cell(r, 14).value = f"=M{r}/L{r}"  # N
        if report.get("general_initial") and report.get("general_subscribed"):
            ws.cell(r, 18).value = f"=Q{r}/P{r}"  # R
        if report.get("esop_initial") and report.get("esop_final") is not None:
            ws.cell(r, 22).value = f"=U{r}/T{r}"  # V

        # 5) 주관/인수 영역 — H025 fallback (기존 셀 모두 빈 경우만)
        lead_col_start = config_ecm.COL_IPO["주관_시작"]
        uw_col_start = lead_col_start + len(config_ecm.LEAD_ECM)
        broker_cells_empty = True
        for ci in range(lead_col_start,
                        uw_col_start + len(config_ecm.UW_ECM)):
            if ws.cell(r, ci).value not in (None, ""):
                broker_cells_empty = False
                break
        if broker_cells_empty and report.get("underwriter_amounts"):
            from collections import defaultdict
            agg = defaultdict(float)
            for ua in report["underwriter_amounts"]:
                alias = parser_ecm.broker_alias(ua.get("name", ""))
                if not alias:
                    continue
                # 보고서 금액은 "원" — 억원으로 변환
                amt_won = ua.get("amount_won") or 0
                if amt_won:
                    agg[alias] += amt_won / 1e8
            placed = 0
            for i, broker in enumerate(config_ecm.LEAD_ECM):
                v = agg.get(broker)
                if v and v > 0:
                    ws.cell(r, lead_col_start + i).value = round(v)
                    placed += 1
            for i, broker in enumerate(config_ecm.UW_ECM):
                v = agg.get(broker)
                if v and v > 0:
                    ws.cell(r, uw_col_start + i).value = round(v)
                    placed += 1
            if placed:
                print(f"    [H025 fallback] {name}: report 인수단 {placed}개 broker 셀 채움")

        rec["rcept_no_report"] = report_f.rcept_no
        finalized += 1
        print(f"  [FIX] {name} (row {r}): report={report_f.rcept_no} 데이터 마감")

    if finalized:
        wb.save(xlsx_path)
    print(f"  [완료] {finalized}건 마감")
    return finalized


def _recheck_kind_for_unfinalized_ipo(xlsx_path: Path,
                                      merged_meta: dict) -> int:
    """**사용자 룰 2026-05-29**: 매 cmd_update 마다 **미마감 IPO** 의 KIND 상장예정일을
    재확인 → A(상장일) 갱신. validator H007/H008 과 동일 의도지만 cmd_update 시
    --only-rcepts 로 신규 딜만 검증되는 한계를 보완 (기존 미마감 딜도 매번 재확인).

    대상 (미마감) 판정 = validator `_is_ipo_finalized` 와 동일:
      - 청약결과 L(기관)/P(일반)/T(우리사주) 최초배정 셀 **전부 비어있음** AND
      - meta record 의 rcept_no_report **없음**
    제외:
      - verified (수동 확정) row — 완전 보호
      - 증권발행실적보고서 확정 상장일이 있는 deal (finalized → report 우선, KIND 무시)

    동작 (validator H007/H008 동일):
      - A '미정'/빈칸 인데 KIND 에 예정일 있으면 → A = KIND 예정일
      - A 가 날짜인데 KIND 현재 예정일과 다르면 → A = KIND 예정일 (drift 반영)
      - KIND 에 예정일 없으면 → 그대로 둠
    A 변경분은 caller 의 A기준 오름차순 재정렬이 처리.

    Returns: A 갱신된 row 수.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    if "IPO" not in wb.sheetnames:
        return 0
    ws = wb["IPO"]

    meta_by_row = {rec.get("row_idx"): rec
                   for rec in merged_meta.get("ipo_records", [])}
    verified_rows = set()
    for rec in merged_meta.get("ipo_records", []):
        if rec.get("verified"):
            verified_rows.add(rec.get("row_idx"))
    for o in merged_meta.get("manual_verified_orphans", []):
        if o.get("verified") and o.get("sheet") == "IPO":
            verified_rows.add(o.get("row_idx"))

    kind_cache: dict = {}

    def _kind_date(corp_name):
        if not corp_name:
            return None
        if corp_name in kind_cache:
            return kind_cache[corp_name]
        try:
            sched = kind_client.fetch_listing_schedule(corp_name)
            out = sched.listing_date_planned if sched else None
        except Exception as e:
            print(f"  [WARN] KIND 조회 실패 ({corp_name}): {e}")
            out = None
        kind_cache[corp_name] = out
        return out

    updated = 0
    checked = 0
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        if r in verified_rows:
            continue
        # 마무리된 IPO skip — 청약결과(L/P/T) 채워졌거나 report 등록됨
        if (ws.cell(r, 12).value is not None
                or ws.cell(r, 16).value is not None
                or ws.cell(r, 20).value is not None):
            continue
        rec = meta_by_row.get(r)
        if rec and rec.get("rcept_no_report"):
            continue
        # 미마감 IPO → KIND 재확인
        checked += 1
        kd = _kind_date(name)
        if kd is None:
            continue
        a_val = ws.cell(r, 1).value
        cur = a_val.date() if hasattr(a_val, "date") else (
            a_val if isinstance(a_val, date) else None)
        if cur == kd:
            continue
        ws.cell(r, 1).value = kd
        updated += 1
        old_disp = cur if cur else (a_val if a_val else "미정")
        print(f"  [KIND] {name} (row {r}): 상장예정일 {old_disp} → {kd}")

    if updated:
        wb.save(xlsx_path)
    print(f"  [KIND 재확인] 미마감 IPO {checked}건 조회 / {updated}건 갱신")
    return updated


def cmd_update(start: date, end: date, xlsx: Optional[Path] = None,
                run_validator: bool = True) -> None:
    """기존 xlsx 보존 + 사용자 기간 신규 ECM 공시만 추가 (DCM cmd_update 패턴).

    동작:
      1. 기존 meta.json 의 모든 rcept_no 모음 = processed_rcepts
      2. list_ecm_filings(start, end) → primary
      3. processed 제외 = new_filings (중복 자동 dedup)
      4. group_into_deals + lazy backfill + process_deal → new_results
      5. 기존 데이터 행 다음부터 _write_*_row append
      6. meta.json 병합 + persist_auto_added
      7. (default) validator_ecm 자동 실행 → Pass 2.5 가 sort + dedup
    """
    print(f"=== ECM 공시 업데이트 {start} ~ {end} ===")
    xlsx = xlsx or config_ecm.ECM_XLSX_TEST
    if not xlsx.exists():
        print(f"  [ERR] {xlsx} 없음. cmd_collect 로 먼저 생성 필요.")
        return

    meta_path = xlsx.with_suffix(".meta.json")
    if not meta_path.exists():
        print(f"  [ERR] {meta_path} 없음. cmd_collect 후 retry.")
        return
    existing_meta = json.loads(meta_path.read_text(encoding="utf-8"))

    processed_rcepts: set[str] = set()
    for rec in existing_meta.get("ipo_records", []):
        for k in ("rcept_no_stage1", "rcept_no_final", "rcept_no_report"):
            v = rec.get(k)
            if v:
                processed_rcepts.add(v)
    for rec in existing_meta.get("rights_records", []):
        for k in ("rcept_no_stage1", "rcept_no_final1", "rcept_no_final2"):
            v = rec.get(k)
            if v:
                processed_rcepts.add(v)
    print(f"  기존: IPO {len(existing_meta.get('ipo_records', []))}건 + "
          f"유증 {len(existing_meta.get('rights_records', []))}건 / "
          f"processed rcept {len(processed_rcepts)}")

    mappings = load_mappings()

    # primary
    print(f"  [1] primary fetch: {start} ~ {end}")
    primary = dart_client.list_ecm_filings(start, end)
    print(f"      ECM 공시 {len(primary)} 건")

    new_primary = [f for f in primary if f.rcept_no not in processed_rcepts]
    print(f"      이미 처리된 {len(primary) - len(new_primary)}건 skip → "
          f"신규 {len(new_primary)}건")

    if not new_primary:
        print("  [완료] 신규 공시 없음")
        return

    # deal grouping
    deals = group_into_deals(new_primary)
    print(f"  [2] 신규 딜 그룹: {len(deals)} 개")

    # lazy backfill (cmd_collect 와 동일 — dropdown 으로 누락 보강)
    needs_bf = []
    for d in deals:
        if d.stage1 is None:
            needs_bf.append(d)
        elif parser_ecm.classify_filing(d.stage1.report_nm) != "stage1":
            needs_bf.append(d)
        elif len(d.finals) == 1:
            needs_bf.append(d)
    if needs_bf:
        print(f"  [3] dropdown 백필 필요: {len(needs_bf)} deal")
        for d in needs_bf:
            sample = d.stage1
            if sample is None:
                for arr in (d.amends, d.finals, d.reports):
                    if arr:
                        sample = arr[0]
                        break
            if sample is None:
                continue
            try:
                dropdown = dart_client.fetch_deal_filings(sample.rcept_no)
            except Exception as e:
                print(f"      [WARN] {d.corp_name} dropdown 실패: {e}")
                continue
            true_stage1 = None
            for f in dropdown:
                if parser_ecm.classify_filing(f.report_nm) == "stage1":
                    true_stage1 = f
                    break
            existing_final_rcepts = {ff.rcept_no for ff in d.finals}
            new_finals = []
            for f in dropdown:
                if f.rcept_no in existing_final_rcepts:
                    continue
                if parser_ecm.classify_filing(f.report_nm) == "final":
                    f.corp_name = d.corp_name
                    f.corp_code = d.corp_code
                    new_finals.append(f)
            if new_finals:
                d.finals.extend(new_finals)
                d.finals.sort(key=lambda x: x.rcept_no)
            if true_stage1 and d.stage1 is None:
                true_stage1.corp_name = d.corp_name
                true_stage1.corp_code = d.corp_code
                d.stage1 = true_stage1

    # process_deal + broker aggregate (cmd_collect 와 동일)
    new_results = []
    print(f"  [4] process_deal × {len(deals)}")
    for deal in deals:
        # 사용자 룰 2026-05-26: 철회된 deal skip
        if deal.is_withdrawn:
            print(f"    [철회 skip] {deal.corp_name} "
                  f"(stage1={deal.stage1.rcept_no if deal.stage1 else 'N/A'})")
            continue
        try:
            res = process_deal(deal)
        except Exception as e:
            print(f"    [ERROR] process_deal 실패 {deal.corp_name}: {e}")
            import traceback; traceback.print_exc()
            continue
        if res is None:
            continue
        if res.kind == "ipo" and res.ipo_record is not None:
            # 사용자 룰 2026-05-25: IPO broker amount = 첫 번째 [발행조건확정] 본문.
            final_rows = getattr(res.ipo_record, "_final_underwriter_rows", [])
            eok, leads, perf = aggregate_broker_amounts(final_rows, mappings=mappings)
            res.underwriter_amounts_eok = eok
            res.lead_aliases = leads
            res.lead_perf = perf
            res.listing_date_planned = getattr(res.ipo_record, "_listing_date", None)
            res.sort_key = res.listing_date_planned
        elif res.kind == "rights" and res.rights_record is not None:
            uw_rows = getattr(res.rights_record, "_final_underwriter_amounts", [])
            eok, leads, perf = aggregate_broker_amounts(uw_rows, mappings=mappings)
            res.underwriter_amounts_eok = eok
            res.lead_aliases = leads
            res.lead_perf = perf
            res.sort_key = res.rights_record.record_date
        new_results.append(res)

    if not new_results:
        print("  [완료] 처리된 신규 deal 없음")
        # **사용자 룰 2026-05-29**: 신규 deal 0 이어도 IPO 미마감 deal 의 보고서
        # follow-up 은 계속 시도 (보고서가 별도 공시로 올라오는 경우 매번 체크).
        merged_meta_only_existing = {
            "ipo_records": list(existing_meta.get("ipo_records", [])),
            "rights_records": list(existing_meta.get("rights_records", [])),
        }
        finalized_n = _finalize_overdue_ipo_reports(xlsx, merged_meta_only_existing)
        kind_n = _recheck_kind_for_unfinalized_ipo(xlsx, merged_meta_only_existing)
        if finalized_n > 0 or kind_n > 0:
            # finalize/KIND 가 A 셀 변경했을 가능성 → 재정렬 + meta row_idx 갱신 + meta 저장
            print(f"\n  [A 기준 오름차순 정렬] (finalize/KIND 후 후속 정리)")
            from validator_ecm import sort_data_rows_by_date
            wb_sort = openpyxl.load_workbook(xlsx)
            for sn in ["IPO", "유상증자"]:
                if sn not in wb_sort.sheetnames:
                    continue
                ws_s = wb_sort[sn]
                row_map = sort_data_rows_by_date(ws_s, ref_row=3, date_col=1)
                key = "ipo_records" if sn == "IPO" else "rights_records"
                moved = 0
                for rec in merged_meta_only_existing.get(key, []):
                    old = rec.get("row_idx")
                    if old in row_map and row_map[old] != old:
                        rec["row_idx"] = row_map[old]
                        moved += 1
                print(f"    {sn}: {moved}개 row 위치 갱신")
            wb_sort.save(xlsx)
            meta_path.write_text(
                json.dumps(merged_meta_only_existing, ensure_ascii=False, indent=2),
                encoding="utf-8")
        return

    # append 모드 — 기존 데이터 보존하고 신규만 추가
    print(f"  [5] xlsx append: {xlsx.name}")
    import excel_writer  # cmd_update 전용 import (module top-level 아님)
    wb = openpyxl.load_workbook(xlsx)
    ws_ipo = wb["IPO"]
    ws_rights = wb["유상증자"]
    excel_writer._expand_broker_columns(
        ws_ipo, leads=config_ecm.LEAD_ECM, uws=config_ecm.UW_ECM)
    excel_writer._expand_broker_columns(
        ws_rights, leads=config_ecm.LEAD_ECM, uws=config_ecm.UW_ECM)

    def _last_data_row(ws) -> int:
        last = 2
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=2).value:
                last = r
        return last

    ipo_results = [r for r in new_results if r.kind == "ipo"]
    rights_results = [r for r in new_results if r.kind == "rights"]
    ipo_results.sort(key=lambda r: r.sort_key or date.max)
    rights_results.sort(key=lambda r: r.sort_key or date.max)

    # **사용자 룰 2026-05-29**: 기존 row 매칭 키를 **stage1 rcept_no** 기반으로.
    # 이전엔 (date, corp_name) 키 사용 → 정정공시로 기준일 변경 시 dedup 실패하고
    # 새 row 추가됨 (한울반도체/형지I&C 등 multi-row 분할 문제). DCM 의 (issuer,
    # series) 처럼 안정적 식별자 사용. stage1 rcept_no 는 정정공시 시에도 불변.
    def _build_existing_keys_by_stage1(meta_records, sheet_name_for_fallback_ws=None):
        keys = {}
        for rec in meta_records:
            s1 = rec.get("rcept_no_stage1")
            ri = rec.get("row_idx")
            if s1 and ri:
                keys[s1] = ri
        return keys

    ipo_existing_keys = _build_existing_keys_by_stage1(
        existing_meta.get("ipo_records", []))
    rights_existing_keys = _build_existing_keys_by_stage1(
        existing_meta.get("rights_records", []))

    def _date_key(d):
        return d.isoformat() if hasattr(d, "isoformat") else str(d or "")

    new_ipo_meta = []
    ipo_appended = ipo_updated = 0
    for dr in ipo_results:
        if dr.ipo_record is None:
            continue
        # stage1 rcept_no 기반 매칭 — 정정공시 시에도 안정적
        key = dr.ipo_record.rcept_no_stage1 or ""
        if key in ipo_existing_keys:
            # 기존 row 업데이트 (덮어쓰기 — 신규 fetch 가 최신 데이터)
            r = ipo_existing_keys[key]
            _write_ipo_row(ws_ipo, r, dr)
            # 기존 meta record 갱신 (rcept_no 등 보충)
            for mrec in existing_meta.get("ipo_records", []):
                if mrec.get("row_idx") == r:
                    for k_, v_ in [("corp_code", dr.ipo_record.corp_code),
                                    ("rcept_no_stage1", dr.ipo_record.rcept_no_stage1),
                                    ("rcept_no_final", dr.ipo_record.rcept_no_final)]:
                        if v_:
                            mrec[k_] = v_
                    break
            ipo_updated += 1
        else:
            # 신규 row 추가
            r = _last_data_row(ws_ipo) + 1
            _write_ipo_row(ws_ipo, r, dr)
            new_ipo_meta.append({
                "row_idx": r,
                "corp_name": dr.ipo_record.issuer or "",
                "corp_code": dr.ipo_record.corp_code or "",
                "rcept_no_stage1": dr.ipo_record.rcept_no_stage1 or "",
                "rcept_no_final": dr.ipo_record.rcept_no_final or "",
            })
            ipo_existing_keys[key] = r  # 다음 신규 deal 의 중복 체크 위함
            ipo_appended += 1

    new_rights_meta = []
    rights_appended = rights_updated = 0
    for dr in rights_results:
        if dr.rights_record is None:
            continue
        # stage1 rcept_no 기반 매칭 — 정정공시 시에도 안정적
        key = dr.rights_record.rcept_no_stage1 or ""
        if key in rights_existing_keys:
            r = rights_existing_keys[key]
            _write_rights_row(ws_rights, r, dr)
            for mrec in existing_meta.get("rights_records", []):
                if mrec.get("row_idx") == r:
                    for k_, v_ in [("corp_code", dr.rights_record.corp_code),
                                    ("rcept_no_stage1", dr.rights_record.rcept_no_stage1),
                                    ("rcept_no_final1", dr.rights_record.rcept_no_final1),
                                    ("rcept_no_final2", dr.rights_record.rcept_no_final2)]:
                        if v_:
                            mrec[k_] = v_
                    break
            rights_updated += 1
        else:
            r = _last_data_row(ws_rights) + 1
            # multi-issue 시 메인 + extra row 모두 작성 (last_r 가 마지막 row)
            last_r = _write_rights_row(ws_rights, r, dr)
            # 메인 + 모든 extra row 각각 meta record 등록 (rcept_no 공유)
            for rr in range(r, last_r + 1):
                new_rights_meta.append({
                    "row_idx": rr,
                    "corp_name": dr.rights_record.issuer or "",
                    "corp_code": dr.rights_record.corp_code or "",
                    "rcept_no_stage1": dr.rights_record.rcept_no_stage1 or "",
                    "rcept_no_final1": dr.rights_record.rcept_no_final1 or "",
                    "rcept_no_final2": dr.rights_record.rcept_no_final2 or "",
                })
            rights_existing_keys[key] = r
            rights_appended += 1
            if last_r > r:
                rights_appended += (last_r - r)  # extra row 도 count

    wb.save(xlsx)
    print(f"  IPO: 신규 {ipo_appended}건 추가, 기존 {ipo_updated}건 갱신 (중복 dedup)")
    print(f"  유증: 신규 {rights_appended}건 추가, 기존 {rights_updated}건 갱신 (중복 dedup)")

    # meta 병합 — 기존 (= 갱신된 record 포함) + new (신규 row)
    merged_meta = {
        "ipo_records": list(existing_meta.get("ipo_records", [])) + new_ipo_meta,
        "rights_records": list(existing_meta.get("rights_records", [])) + new_rights_meta,
    }

    # **사용자 룰 2026-05-28**: cmd_update 는 이번 기간 공시와 연관된 deal 만 처리.
    # 이전엔 _finalize_overdue_rights_deals 호출로 납입일 도래한 무관 deal 의
    # O(최종_가액) 자동 patch 했지만, 이는 무관 deal 데이터 변경이라 룰 위반.
    # 영구 비활성. 필요 시 별도 cmd_finalize_overdue 명령으로 분리 가능.

    # **사용자 룰 2026-05-29 (IPO 전용 예외)**: [발행조건확정] 까지 도달했으나
    # 증권발행실적보고서 미수집 IPO 딜에 한해 자동 follow-up. 상장일은 보고서값으로
    # 덮어쓰기 (옵션 A). 변경된 A 셀은 아래 A-sort 가 row 위치를 재정렬함.
    _finalize_overdue_ipo_reports(xlsx, merged_meta)
    # **사용자 룰 2026-05-29**: 매 cmd_update 마다 미마감 IPO 의 KIND 상장예정일 재확인.
    _recheck_kind_for_unfinalized_ipo(xlsx, merged_meta)

    # **사용자 룰 2026-05-28**: 신규 row append 후 IPO/유증 시트 A(기준일/상장일)
    # 오름차순 정렬만 적용. validator 의 Pass 2.4/2.5/2.6 (dedup/sort/수식 재작성)
    # 모두 차단. self-row 수식 자동 갱신은 sort_data_rows_by_date 함수 자체에
    # 포함됨 (=E48/F48 → =E13/F13 단순 행번호 치환, 데이터 손상 X).
    print(f"\n  [A 기준 오름차순 정렬]")
    from validator_ecm import sort_data_rows_by_date
    wb_sort = openpyxl.load_workbook(xlsx)
    for sn in ["IPO", "유상증자"]:
        if sn not in wb_sort.sheetnames:
            continue
        ws_s = wb_sort[sn]
        row_map = sort_data_rows_by_date(ws_s, ref_row=3, date_col=1)
        # meta row_idx 갱신
        key = "ipo_records" if sn == "IPO" else "rights_records"
        moved = 0
        for rec in merged_meta.get(key, []):
            old = rec.get("row_idx")
            if old in row_map and row_map[old] != old:
                rec["row_idx"] = row_map[old]
                moved += 1
        print(f"    {sn}: {moved}개 row 위치 갱신")
    wb_sort.save(xlsx)

    meta_path.write_text(
        json.dumps(merged_meta, ensure_ascii=False, indent=2), encoding="utf-8")

    persist_auto_added(mappings)
    print(f"  [완료] 신규 IPO {len(new_ipo_meta)} + 유증 {len(new_rights_meta)} 추가 → "
          f"총 IPO {len(merged_meta['ipo_records'])} + 유증 {len(merged_meta['rights_records'])}")

    # validator 자동 실행 — 사용자 룰 2026-05-25: **신규 record 만 Layer 1 검증**
    # (이전에 검증 통과한 누적 데이터 재검증 안 함, 효율). Pass 2.4 dedup / 2.5 sort 는 전체 유지.
    if run_validator:
        print(f"\n=== validator 자동 실행 ===")
        new_rcepts: list[str] = []
        for rec in new_ipo_meta:
            if rec.get("rcept_no_stage1"):
                new_rcepts.append(rec["rcept_no_stage1"])
        for rec in new_rights_meta:
            if rec.get("rcept_no_stage1"):
                new_rcepts.append(rec["rcept_no_stage1"])

        cmd = [sys.executable, "-X", "utf8", str(config.ROOT / "validator_ecm.py"),
               "--xlsx", str(xlsx)]
        if new_rcepts:
            cmd += ["--only-rcepts", ",".join(new_rcepts)]
            print(f"  (신규 {len(new_rcepts)} 건 대상 검증)")

        import subprocess
        subprocess.run(cmd, check=False, encoding="utf-8")


def cmd_status() -> None:
    """현재 ECM Table test.xlsx + meta 상태 점검."""
    p = config_ecm.ECM_XLSX_TEST
    print(f"파일: {p} (exists={p.exists()})")
    if p.exists():
        wb = openpyxl.load_workbook(p)
        for sn in wb.sheetnames:
            ws = wb[sn]
            data_rows = sum(1 for r in range(3, ws.max_row + 1)
                            if ws.cell(row=r, column=2).value)
            print(f"  sheet {sn!r}: {data_rows} 데이터 행")


def cmd_capture_verified(xlsx: Optional[Path] = None,
                         source: str = "", note: str = "") -> None:
    """**사용자 룰 2026-05-29**: 사용자가 직접 확인·입력한 셀을 baseline 과 diff 해서
    meta 의 record 에 `verified` 로 기록 (완전 보호 플래그 — cmd_update finalize +
    validator 가 해당 row skip).

    baseline: 'ECM Table.captureBaseline.xlsx' (없으면 'ECM Table.preManual_20260529.xlsx').
    diff 후 captureBaseline 을 현재 상태로 갱신 → 다음 캡처는 새 편집분만 잡음.

    verified = {date, source_rcept, note, changed_cells:[{col,letter,header,old,new}]}
    키는 row_idx 로 매칭하되 기록은 record (rcept_no_stage1 anchor) 에 부착 → 정렬돼도 유지.
    """
    import shutil
    from copy import copy as _copy  # noqa
    from openpyxl.utils import get_column_letter

    xlsx = xlsx or config_ecm.ECM_XLSX
    root = xlsx.parent
    meta_path = xlsx.with_suffix(".meta.json")
    rolling = root / "ECM Table.captureBaseline.xlsx"
    if not rolling.exists():
        # baseline 없음 → 현재 상태를 기준선으로 부트스트랩 (이후 편집분부터 감지).
        # 별도 원본 백업에 의존하지 않음 — captureBaseline 만으로 자급.
        shutil.copy2(xlsx, rolling)
        print(f"  captureBaseline 없음 → 현재 상태로 초기화. 이후 수동 편집분부터 캡처됩니다.")
        return
    baseline = rolling
    print(f"=== 수동 편집 캡처 (baseline = {baseline.name}) ===")

    wb_base = openpyxl.load_workbook(baseline, data_only=False)
    wb_cur = openpyxl.load_workbook(xlsx, data_only=False)
    meta = json.loads(meta_path.read_text(encoding="utf-8"))

    def norm(v):
        if v is None or v == "":
            return None
        return v

    TODAY = date.today().isoformat()
    total_changes = 0
    unanchored = []  # meta record 없는 변경 row
    report = []

    for sheet, meta_key in [("IPO", "ipo_records"), ("유상증자", "rights_records")]:
        if sheet not in wb_cur.sheetnames or sheet not in wb_base.sheetnames:
            continue
        ws_b = wb_base[sheet]
        ws_c = wb_cur[sheet]
        recs_by_row = {rec.get("row_idx"): rec for rec in meta.get(meta_key, [])}
        max_r = max(ws_b.max_row, ws_c.max_row)
        max_c = max(ws_b.max_column, ws_c.max_column)
        for r in range(3, max_r + 1):
            name_b = ws_b.cell(r, 2).value
            name_c = ws_c.cell(r, 2).value
            # 정렬·삽입 가드 — 같은 row_idx 의 회사명이 다르면 행이 밀린 것
            if norm(name_b) != norm(name_c) and name_b and name_c:
                report.append(f"  [WARN] {sheet} row {r}: 회사명 baseline≠현재 "
                              f"('{name_b}'→'{name_c}') — 행 이동 의심, 이 행 skip")
                continue
            changed = []
            for c in range(1, max_c + 1):
                vb = norm(ws_b.cell(r, c).value)
                vc = norm(ws_c.cell(r, c).value)
                if vb != vc:
                    changed.append({
                        "col": c,
                        "letter": get_column_letter(c),
                        "header": ws_c.cell(2, c).value,
                        "old": vb if not hasattr(vb, "isoformat") else vb.isoformat(),
                        "new": vc if not hasattr(vc, "isoformat") else vc.isoformat(),
                    })
            if not changed:
                continue
            total_changes += len(changed)

            def _merge_verified(container_get, container_set):
                v = container_get() or {"date": TODAY, "source_rcept": source,
                                        "note": note or "사용자 직접 확인 입력",
                                        "changed_cells": []}
                by_col = {cc["col"]: cc for cc in v["changed_cells"]}
                for cc in changed:
                    by_col[cc["col"]] = cc
                v["changed_cells"] = [by_col[k] for k in sorted(by_col)]
                v["date"] = TODAY
                if source:
                    v["source_rcept"] = source
                if note:
                    v["note"] = note
                container_set(v)

            rec = recs_by_row.get(r)
            cells_str = ", ".join(f"{cc['letter']}({cc['header']})={cc['new']}"
                                  for cc in changed)
            if rec:
                _merge_verified(lambda: rec.get("verified"),
                                lambda v: rec.__setitem__("verified", v))
                report.append(f"  [VERIFIED] {sheet} row {r} {name_c}: {cells_str}")
            else:
                # 메타 record 없는 row (구 데이터 orphan / multi-issue 추가행) →
                # 별도 ledger 에 verified 저장. 보호 로직이 동일하게 skip.
                orphans = meta.setdefault("manual_verified_orphans", [])
                o = next((x for x in orphans
                          if x.get("sheet") == sheet and x.get("row_idx") == r), None)
                if o is None:
                    o = {"sheet": sheet, "row_idx": r, "corp_name": name_c}
                    orphans.append(o)
                o["corp_name"] = name_c
                _merge_verified(lambda: o.get("verified"),
                                lambda v: o.__setitem__("verified", v))
                unanchored.append((sheet, r, name_c, changed))
                report.append(f"  [VERIFIED-orphan] {sheet} row {r} {name_c}: {cells_str}")

    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2),
                         encoding="utf-8")
    # 다음 캡처용 baseline 갱신
    shutil.copy2(xlsx, rolling)

    print("\n".join(report) if report else "  변경 없음.")
    print(f"\n  총 {total_changes}개 셀 변경 캡처 / verified 기록 완료")
    if unanchored:
        print(f"  [주의] meta record 없는 {len(unanchored)}개 row 는 "
              f"orphan ledger 에 verified 저장 (보호 동일 적용)")
    print(f"  baseline 갱신: {rolling.name} (다음 캡처는 이후 편집분만 잡음)")


def parse_iso_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def main():
    ap = argparse.ArgumentParser(prog="main_ecm",
                                  description="ECM 지분증권 공시 수집 + 파싱.")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_col = sub.add_parser("collect", help="기간 내 ECM 공시 수집 (기존 데이터 reset)")
    p_col.add_argument("start", help="YYYY-MM-DD")
    p_col.add_argument("end", help="YYYY-MM-DD")
    p_col.add_argument("--corp-code", default="", help="특정 corp 만 (refetch)")
    p_col.add_argument("--xlsx", default=None, type=Path,
                       help="대상 xlsx 경로 (default = ECM Table test.xlsx)")

    p_upd = sub.add_parser("update", help="기존 xlsx 보존 + 신규 기간 공시만 추가")
    p_upd.add_argument("start", help="YYYY-MM-DD")
    p_upd.add_argument("end", help="YYYY-MM-DD")
    p_upd.add_argument("--xlsx", default=None, type=Path,
                       help="대상 xlsx 경로 (default = ECM Table test.xlsx)")
    p_upd.add_argument("--no-validator", action="store_true",
                       help="update 후 validator_ecm 자동 실행 안 함")

    sub.add_parser("status", help="현재 테스트 xlsx 상태")

    p_cap = sub.add_parser("capture-verified",
                           help="사용자 수동 편집을 baseline diff 로 meta 에 verified 기록")
    p_cap.add_argument("--xlsx", default=None, type=Path,
                       help="대상 xlsx 경로 (default = ECM Table.xlsx)")
    p_cap.add_argument("--source", default="",
                       help="출처 공시번호 (이번 캡처 전체에 적용, 선택)")
    p_cap.add_argument("--note", default="",
                       help="메모 (선택)")

    args = ap.parse_args()
    if args.cmd == "collect":
        cmd_collect(parse_iso_date(args.start), parse_iso_date(args.end),
                    corp_code=args.corp_code, xlsx=args.xlsx)
    elif args.cmd == "update":
        cmd_update(parse_iso_date(args.start), parse_iso_date(args.end),
                   xlsx=args.xlsx, run_validator=not args.no_validator)
    elif args.cmd == "status":
        cmd_status()
    elif args.cmd == "capture-verified":
        cmd_capture_verified(xlsx=args.xlsx, source=args.source, note=args.note)


if __name__ == "__main__":
    main()
